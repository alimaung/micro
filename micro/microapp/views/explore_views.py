"""
Roll management views for the microapp.
These views handle CRUD operations for film rolls.
"""

from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.views.decorators.csrf import csrf_exempt
import json
from django.views.decorators.http import require_http_methods

from ..models import Roll, Project

@csrf_exempt
@login_required
def create_roll(request):
    """
    API endpoint to create a new roll.
    
    Args:
        Various roll attributes in JSON request body
        
    Returns:
        JSON response with roll ID and status
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Get the project (ensure user owns it)
            project_id = data.get('project_id')
            if not project_id:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Project ID is required'
                }, status=400)
            
            try:
                project = Project.objects.get(id=project_id, owner=request.user)
            except Project.DoesNotExist:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Project not found or access denied'
                }, status=404)
            
            # Create a new roll
            roll = Roll(
                project=project,
                roll_id=data.get('roll_id'),
                film_number=data.get('film_number', ''),
                film_type=data.get('film_type', '16mm'),
                capacity=data.get('capacity', 2940),
                pages_used=data.get('pages_used', 0),
                pages_remaining=data.get('pages_remaining'),
                status=data.get('status', 'active'),
                has_split_documents=data.get('has_split_documents', False),
                is_partial=data.get('is_partial', False),
                remaining_capacity=data.get('remaining_capacity', 0),
                usable_capacity=data.get('usable_capacity', 0),
                film_number_source=data.get('film_number_source', '')
            )
            
            # Calculate pages_remaining if not provided
            if roll.pages_remaining is None:
                roll.pages_remaining = roll.capacity - roll.pages_used
            
            roll.save()
            
            return JsonResponse({
                'status': 'success',
                'roll_id': roll.id,
                'message': f'Roll {roll.roll_id} created successfully'
            })
            
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=400)
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@login_required
def list_rolls(request):
    """
    API endpoint to list all rolls with filtering options.
    
    Args:
        Various filter parameters as GET parameters
        
    Returns:
        JSON response with paginated roll list
    """
    try:
        # Get query parameters for filtering
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 20))
        
        # Initialize queryset with user's rolls (through project ownership)
        queryset = Roll.objects.filter(project__owner=request.user)
        
        # Apply filters if provided
        project_id = request.GET.get('project_id')
        if project_id:
            queryset = queryset.filter(project_id=project_id)
            
        film_number = request.GET.get('film_number')
        if film_number:
            queryset = queryset.filter(film_number__icontains=film_number)
            
        film_type = request.GET.get('film_type')
        if film_type:
            queryset = queryset.filter(film_type__iexact=film_type)
            
        status = request.GET.get('status')
        if status:
            queryset = queryset.filter(status__iexact=status)
            
        has_split_documents = request.GET.get('has_split_documents')
        if has_split_documents is not None:
            has_split_bool = has_split_documents.lower() in ['true', '1', 't', 'y', 'yes']
            queryset = queryset.filter(has_split_documents=has_split_bool)
            
        is_partial = request.GET.get('is_partial')
        if is_partial is not None:
            is_partial_bool = is_partial.lower() in ['true', '1', 't', 'y', 'yes']
            queryset = queryset.filter(is_partial=is_partial_bool)
            
        # Date range filtering
        date_from = request.GET.get('date_from')
        if date_from:
            queryset = queryset.filter(creation_date__gte=date_from)
            
        date_to = request.GET.get('date_to')
        if date_to:
            queryset = queryset.filter(creation_date__lte=date_to)
            
        # Search term
        search_term = request.GET.get('search')
        if search_term:
            queryset = queryset.filter(
                Q(film_number__icontains=search_term) | 
                Q(project__archive_id__icontains=search_term) |
                Q(status__icontains=search_term) |
                Q(film_type__icontains=search_term)
            )
            
        # Sorting
        sort_field = request.GET.get('sort_field', 'creation_date')
        sort_order = request.GET.get('sort_order', 'desc')
        
        # Map sort field to model field
        field_mapping = {
            'id': 'id',
            'roll_id': 'roll_id',
            'project_id': 'project_id',
            'project_archive_id': 'project__archive_id',
            'film_number': 'film_number',
            'film_type': 'film_type',
            'capacity': 'capacity',
            'pages_used': 'pages_used',
            'pages_remaining': 'pages_remaining',
            'status': 'status',
            'has_split_documents': 'has_split_documents',
            'is_partial': 'is_partial',
            'remaining_capacity': 'remaining_capacity',
            'usable_capacity': 'usable_capacity',
            'film_number_source': 'film_number_source',
            'creation_date': 'creation_date',
            'utilization': 'pages_used',  # We'll calculate this in frontend
        }
        
        # Get the actual model field name, default to creation_date if not found
        actual_sort_field = field_mapping.get(sort_field, 'creation_date')
        
        # Apply sort order (desc = descending, asc = ascending)
        if sort_order.lower() == 'desc':
            actual_sort_field = f'-{actual_sort_field}'
            
        queryset = queryset.order_by(actual_sort_field)
        
        # Count total records (for pagination)
        total_records = queryset.count()
        total_pages = (total_records + page_size - 1) // page_size  # Ceiling division
        
        # Apply pagination
        start = (page - 1) * page_size
        end = start + page_size
        paginated_queryset = queryset[start:end]
        
        # Serialize paginated queryset
        rolls_data = []
        for roll in paginated_queryset:
            # Calculate utilization percentage
            utilization = (roll.pages_used / roll.capacity * 100) if roll.capacity > 0 else 0
            
            rolls_data.append({
                'id': roll.id,
                'roll_id': roll.roll_id,
                'project_id': roll.project.id,
                'project_archive_id': roll.project.archive_id,
                'film_number': roll.film_number or '',
                'film_type': roll.film_type,
                'capacity': roll.capacity,
                'pages_used': roll.pages_used,
                'pages_remaining': roll.pages_remaining,
                'status': roll.status,
                'has_split_documents': roll.has_split_documents,
                'is_partial': roll.is_partial,
                'remaining_capacity': roll.remaining_capacity,
                'usable_capacity': roll.usable_capacity,
                'film_number_source': roll.film_number_source or '',
                'creation_date': roll.creation_date.strftime('%Y-%m-%d'),
                'utilization': round(utilization, 1),
                'is_full': roll.is_full,
                # Additional fields for UI compatibility
                'project_name': roll.project.archive_id,
                'date_created': roll.creation_date.strftime('%Y-%m-%d'),
            })
        
        return JsonResponse({
            'status': 'success',
            'total': total_records,
            'page': page,
            'page_size': page_size,
            'total_pages': total_pages,
            'results': rolls_data
        })
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)

@login_required
def get_roll(request, roll_id):
    """
    API endpoint to get a single roll by ID.
    
    Args:
        roll_id: Roll ID (URL parameter)
        
    Returns:
        JSON response with roll details
    """
    try:
        roll = Roll.objects.get(id=roll_id, project__owner=request.user)
        
        # Calculate utilization percentage
        utilization = (roll.pages_used / roll.capacity * 100) if roll.capacity > 0 else 0
        
        roll_data = {
            'id': roll.id,
            'roll_id': roll.roll_id,
            'project_id': roll.project.id,
            'project_archive_id': roll.project.archive_id,
            'film_number': roll.film_number or '',
            'film_type': roll.film_type,
            'capacity': roll.capacity,
            'pages_used': roll.pages_used,
            'pages_remaining': roll.pages_remaining,
            'status': roll.status,
            'has_split_documents': roll.has_split_documents,
            'is_partial': roll.is_partial,
            'remaining_capacity': roll.remaining_capacity,
            'usable_capacity': roll.usable_capacity,
            'film_number_source': roll.film_number_source or '',
            'creation_date': roll.creation_date.strftime('%Y-%m-%d'),
            'utilization': round(utilization, 1),
            'is_full': roll.is_full,
            # Additional fields for UI compatibility
            'project_name': roll.project.archive_id,
            'date_created': roll.creation_date.strftime('%Y-%m-%d'),
        }
        
        return JsonResponse({
            'status': 'success',
            'roll': roll_data
        })
        
    except Roll.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'Roll not found'
        }, status=404)
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)

@csrf_exempt
@login_required
def update_roll(request, roll_id):
    """
    API endpoint to update a roll.
    
    Args:
        roll_id: Roll ID (URL parameter)
        Various roll attributes in JSON request body
        
    Returns:
        JSON response with status and message
    """
    if request.method == 'PUT':
        try:
            roll = Roll.objects.get(id=roll_id, project__owner=request.user)
            data = json.loads(request.body)
            
            # Update roll fields if provided in the request
            if 'roll_id' in data:
                roll.roll_id = data['roll_id']
            if 'film_number' in data:
                roll.film_number = data['film_number']
            if 'film_type' in data:
                roll.film_type = data['film_type']
            if 'capacity' in data:
                roll.capacity = data['capacity']
            if 'pages_used' in data:
                roll.pages_used = data['pages_used']
            if 'pages_remaining' in data:
                roll.pages_remaining = data['pages_remaining']
            if 'status' in data:
                roll.status = data['status']
            if 'has_split_documents' in data:
                roll.has_split_documents = data['has_split_documents']
            if 'is_partial' in data:
                roll.is_partial = data['is_partial']
            if 'remaining_capacity' in data:
                roll.remaining_capacity = data['remaining_capacity']
            if 'usable_capacity' in data:
                roll.usable_capacity = data['usable_capacity']
            if 'film_number_source' in data:
                roll.film_number_source = data['film_number_source']
                
            # Recalculate pages_remaining if capacity or pages_used changed
            if 'capacity' in data or 'pages_used' in data:
                roll.pages_remaining = roll.capacity - roll.pages_used
                
            # Save the updated roll
            roll.save()
            
            return JsonResponse({
                'status': 'success',
                'message': f'Roll {roll.roll_id} updated successfully'
            })
            
        except Roll.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'message': 'Roll not found'
            }, status=404)
            
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=400)
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@csrf_exempt
@login_required
def delete_roll(request, roll_id):
    """
    API endpoint to delete a roll.
    
    Args:
        roll_id: Roll ID (URL parameter)
        
    Returns:
        JSON response with status and message
    """
    if request.method == 'DELETE':
        try:
            roll = Roll.objects.get(id=roll_id, project__owner=request.user)
            roll_identifier = f"Roll {roll.roll_id} ({roll.film_number or 'No film number'})"
            roll.delete()
            
            return JsonResponse({
                'status': 'success',
                'message': f'{roll_identifier} deleted successfully'
            })
            
        except Roll.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'message': 'Roll not found'
            }, status=404)
            
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=400)
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@login_required
@require_http_methods(["POST"])
def export_rolls_excel(request):
    """
    API endpoint to export rolls as Excel file.
    
    Args:
        request: HTTP request object with JSON data
        
    Returns:
        Excel file as response
    """
    try:
        import json
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        
        # Parse JSON data from request
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({
                "status": "error",
                "message": "Invalid JSON data"
            }, status=400)
        
        rolls = data.get('rolls', [])
        fields = data.get('fields', [])
        field_names = data.get('fieldNames', [])
        options = data.get('options', {})
        
        if not rolls:
            return JsonResponse({
                "status": "error",
                "message": "No rolls data provided"
            }, status=400)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rolls"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add headers
        for col_idx, field_name in enumerate(field_names, 1):
            cell = ws.cell(row=1, column=col_idx, value=field_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data rows
        for row_idx, roll in enumerate(rolls, 2):
            for col_idx, field in enumerate(fields, 1):
                value = roll.get(field, '')
                
                # Convert boolean values to Yes/No
                if isinstance(value, bool):
                    value = 'Yes' if value else 'No'
                
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for col_idx in range(1, len(fields) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            
            for row in ws[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add auto-filter if requested
        if options.get('autoFilter', True):
            ws.auto_filter.ref = ws.dimensions
        
        # Freeze header row if requested
        if options.get('freezeHeader', True):
            ws.freeze_panes = 'A2'
        
        # Add summary statistics if requested
        if options.get('includeStats', False):
            # Add a new worksheet for statistics
            stats_ws = wb.create_sheet("Statistics")
            
            stats_data = [
                ["Statistic", "Value"],
                ["Total Rolls", len(rolls)],
                ["16mm Rolls", sum(1 for r in rolls if r.get('film_type') == '16mm')],
                ["35mm Rolls", sum(1 for r in rolls if r.get('film_type') == '35mm')],
                ["Full Rolls", sum(1 for r in rolls if r.get('is_full'))],
                ["Partial Rolls", sum(1 for r in rolls if r.get('is_partial'))],
                ["Rolls with Split Documents", sum(1 for r in rolls if r.get('has_split_documents'))],
                ["Total Capacity", sum(r.get('capacity', 0) for r in rolls)],
                ["Total Pages Used", sum(r.get('pages_used', 0) for r in rolls)],
                ["Total Pages Remaining", sum(r.get('pages_remaining', 0) for r in rolls)],
                ["Average Utilization", f"{sum(r.get('utilization', 0) for r in rolls) / len(rolls):.1f}%" if rolls else "0%"],
            ]
            
            for row_idx, (stat, value) in enumerate(stats_data, 1):
                stats_ws.cell(row=row_idx, column=1, value=stat)
                stats_ws.cell(row=row_idx, column=2, value=value)
                
                if row_idx == 1:  # Header row
                    stats_ws.cell(row=row_idx, column=1).font = header_font
                    stats_ws.cell(row=row_idx, column=1).fill = header_fill
                    stats_ws.cell(row=row_idx, column=2).font = header_font
                    stats_ws.cell(row=row_idx, column=2).fill = header_fill
            
            # Auto-adjust column widths for stats
            for col_idx in range(1, 3):
                column_letter = get_column_letter(col_idx)
                max_length = max(len(str(cell.value)) for cell in stats_ws[column_letter] if cell.value)
                stats_ws.column_dimensions[column_letter].width = max_length + 2
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="rolls_export.xlsx"'
        
        return response
        
    except Exception as e:
        return JsonResponse({
            "status": "error",
            "message": f"Failed to export as Excel: {str(e)}"
        }, status=500)

@login_required
@require_http_methods(["POST"])
def export_rolls_pdf(request):
    """
    API endpoint to export rolls as PDF file.
    
    Args:
        request: HTTP request object with JSON data
        
    Returns:
        PDF file as response
    """
    try:
        import json
        from io import BytesIO
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4, legal, landscape, portrait
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from django.http import HttpResponse
        from datetime import datetime
        
        # Parse JSON data from request
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({
                "status": "error",
                "message": "Invalid JSON data"
            }, status=400)
        
        rolls = data.get('rolls', [])
        fields = data.get('fields', [])
        field_names = data.get('fieldNames', [])
        options = data.get('options', {})
        
        if not rolls:
            return JsonResponse({
                "status": "error",
                "message": "No rolls data provided"
            }, status=400)
        
        # Create PDF
        output = BytesIO()
        
        # Set page size and orientation
        page_size = letter  # Default
        if options.get('pageSize') == 'a4':
            page_size = A4
        elif options.get('pageSize') == 'legal':
            page_size = legal
        
        if options.get('orientation') == 'landscape':
            page_size = landscape(page_size)
        else:
            page_size = portrait(page_size)
        
        doc = SimpleDocTemplate(output, pagesize=page_size, topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        # Build story
        story = []
        styles = getSampleStyleSheet()
        
        # Add header if requested
        if options.get('includeHeaderFooter', True):
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            story.append(Paragraph("Rolls Export Report", title_style))
            story.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
            story.append(Spacer(1, 20))
        
        # Prepare table data
        table_data = [field_names]  # Header row
        
        for roll in rolls:
            row = []
            for field in fields:
                value = roll.get(field, '')
                
                # Convert boolean values to Yes/No
                if isinstance(value, bool):
                    value = 'Yes' if value else 'No'
                
                # Truncate long text for PDF
                if isinstance(value, str) and len(value) > 50:
                    value = value[:47] + "..."
                
                row.append(str(value))
            table_data.append(row)
        
        # Create table
        table = Table(table_data)
        
        # Style the table
        table_style = TableStyle([
            # Header row styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Data rows styling
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            
            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
        ])
        
        table.setStyle(table_style)
        story.append(table)
        
        # Add footer if requested
        if options.get('includeHeaderFooter', True):
            story.append(Spacer(1, 20))
            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=8,
                alignment=1  # Center alignment
            )
            story.append(Paragraph(f"Total Rolls: {len(rolls)}", footer_style))
            story.append(Paragraph("Microfilm Management System", footer_style))
        
        # Build PDF
        doc.build(story)
        output.seek(0)
        
        # Create response
        response = HttpResponse(output.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="rolls_export.pdf"'
        
        return response
        
    except Exception as e:
        return JsonResponse({
            "status": "error",
            "message": f"Failed to export as PDF: {str(e)}"
        }, status=500) 