"""
Project management views for the microapp.
These views handle CRUD operations for projects.
"""

from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.views.decorators.csrf import csrf_exempt
import json
from django.views.decorators.http import require_http_methods

from ..models import Project

@csrf_exempt
@login_required
def create_project(request):
    """
    API endpoint to create a new project.
    
    Args:
        Various project attributes in JSON request body
        
    Returns:
        JSON response with project ID and status
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Create a new project with the basic information from transfer
            project = Project(
                archive_id=data.get('archive_id'),
                location=data.get('location'),
                doc_type=data.get('doc_type', ''),
                project_path=data.get('project_path'),
                project_folder_name=data.get('project_folder_name'),
                comlist_path=data.get('comlist_path'),
                output_dir=data.get('output_dir', ''),
                retain_sources=data.get('retain_sources', True),
                add_to_database=data.get('add_to_database', True),
                pdf_folder_path=data.get('pdf_folder_path', ''),
                has_pdf_folder=data.get('has_pdf_folder', False),
                owner=request.user
            )
            project.save()
            
            return JsonResponse({
                'status': 'success',
                'project_id': project.id,
                'message': f'Project {project.archive_id} created successfully'
            })
            
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=400)
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@login_required
def list_projects(request):
    """
    API endpoint to list all projects with filtering options.
    
    Args:
        Various filter parameters as GET parameters
        
    Returns:
        JSON response with paginated project list
    """
    try:
        # Get query parameters for filtering
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 20))
        
        # Initialize queryset with user's projects
        queryset = Project.objects.filter(owner=request.user)
        
        # Apply filters if provided
        archive_id = request.GET.get('archive_id')
        if archive_id:
            queryset = queryset.filter(archive_id__icontains=archive_id)
            
        location = request.GET.get('location')
        if location:
            queryset = queryset.filter(location__iexact=location)
            
        doc_type = request.GET.get('doc_type')
        if doc_type:
            queryset = queryset.filter(doc_type__iexact=doc_type)
            
        has_oversized = request.GET.get('has_oversized')
        if has_oversized is not None:
            has_oversized_bool = has_oversized.lower() in ['true', '1', 't', 'y', 'yes']
            queryset = queryset.filter(has_oversized=has_oversized_bool)
            
        # Date range filtering
        date_from = request.GET.get('date_from')
        if date_from:
            queryset = queryset.filter(created_at__gte=date_from)
            
        date_to = request.GET.get('date_to')
        if date_to:
            queryset = queryset.filter(created_at__lte=date_to)
            
        # Search term
        search_term = request.GET.get('search')
        if search_term:
            queryset = queryset.filter(
                Q(archive_id__icontains=search_term) | 
                Q(project_folder_name__icontains=search_term) |
                Q(doc_type__icontains=search_term) |
                Q(location__icontains=search_term)
            )
            
        # Sorting
        sort_field = request.GET.get('sort_field', 'created_at')
        sort_order = request.GET.get('sort_order', 'desc')
        
        # Debug logging
        print(f"Backend received sort_field: {sort_field}, sort_order: {sort_order}")
        
        # Map sort field to model field - Updated to include all sortable fields
        field_mapping = {
            'id': 'id',
            'project_id': 'id',  # Alternative name for ID
            'archive_id': 'archive_id',
            'location': 'location',
            'doc_type': 'doc_type',
            'project_path': 'project_path',
            'project_folder_name': 'project_folder_name',
            'pdf_folder_path': 'pdf_folder_path',
            'comlist_path': 'comlist_path',
            'output_dir': 'output_dir',
            'has_pdf_folder': 'has_pdf_folder',
            'processing_complete': 'processing_complete',
            'retain_sources': 'retain_sources',
            'add_to_database': 'add_to_database',
            'has_oversized': 'has_oversized',
            'total_pages': 'total_pages',
            'total_pages_with_refs': 'total_pages_with_refs',
            'documents_with_oversized': 'documents_with_oversized',
            'total_oversized': 'total_oversized',
            'created_at': 'created_at',
            'updated_at': 'updated_at',
            'date_created': 'created_at',  # Alternative name for created_at
            'owner': 'owner__username',  # Sort by username for owner field
            'index_path': 'comlist_path',  # Alternative name for comlist_path
            'data_dir': 'output_dir',  # Alternative name for output_dir
            'film_allocation_complete': 'film_allocation_complete',
            'distribution_complete': 'distribution_complete',
            # Legacy field mappings for compatibility
            'created': 'created_at',
            'name': 'project_folder_name',
            'pages': 'total_pages'
        }
        
        # Get the actual model field name, default to created_at if not found
        actual_sort_field = field_mapping.get(sort_field, 'created_at')
        
        # Debug logging
        print(f"Mapped to actual_sort_field: {actual_sort_field}")
        
        # Apply sort order (desc = descending, asc = ascending)
        if sort_order.lower() == 'desc':
            actual_sort_field = f'-{actual_sort_field}'
            
        print(f"Final sort field with order: {actual_sort_field}")
            
        queryset = queryset.order_by(actual_sort_field)
        
        # Count total records (for pagination)
        total_records = queryset.count()
        total_pages = (total_records + page_size - 1) // page_size  # Ceiling division
        
        # Apply pagination
        start = (page - 1) * page_size
        end = start + page_size
        paginated_queryset = queryset[start:end]
        
        # Serialize paginated queryset
        projects_data = []
        for project in paginated_queryset:
            projects_data.append({
                'id': project.id,
                'project_id': project.id,  # For compatibility with the UI
                'archive_id': project.archive_id,
                'location': project.location,
                'doc_type': project.doc_type,
                'project_path': project.project_path,
                'project_folder_name': project.project_folder_name,
                'comlist_path': project.comlist_path,
                'output_dir': project.output_dir,
                'has_pdf_folder': project.has_pdf_folder,
                'processing_complete': project.processing_complete,
                'film_allocation_complete': project.film_allocation_complete,
                'retain_sources': project.retain_sources,
                'add_to_database': project.add_to_database,
                'has_oversized': project.has_oversized,
                'total_pages': project.total_pages or 0,
                'total_pages_with_refs': project.total_pages_with_refs or 0,
                'date_created': project.created_at.strftime('%Y-%m-%d'),
                'updated_at': project.updated_at.strftime('%Y-%m-%d'),
                'owner': project.owner.username,
                # Additional fields for UI compatibility
                'folderName': project.project_folder_name,
                'path': project.project_path,
                'oversized': project.has_oversized,
                'data_dir': project.output_dir,
                'index_path': project.comlist_path
            })
        
        return JsonResponse({
            'status': 'success',
            'total': total_records,
            'page': page,
            'page_size': page_size,
            'total_pages': total_pages,
            'results': projects_data
        })
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)

@login_required
def get_project(request, project_id):
    """
    API endpoint to get a single project by ID.
    
    Args:
        project_id: Project ID (URL parameter)
        
    Returns:
        JSON response with project details
    """
    try:
        project = Project.objects.get(id=project_id, owner=request.user)
        
        project_data = {
            'id': project.id,
            'project_id': project.id,  # For compatibility with the UI
            'archive_id': project.archive_id,
            'location': project.location,
            'doc_type': project.doc_type,
            'project_path': project.project_path,
            'project_folder_name': project.project_folder_name,
            'comlist_path': project.comlist_path,
            'output_dir': project.output_dir,
            'has_pdf_folder': project.has_pdf_folder,
            'processing_complete': project.processing_complete,
            'film_allocation_complete': project.film_allocation_complete,
            'retain_sources': project.retain_sources,
            'add_to_database': project.add_to_database,
            'has_oversized': project.has_oversized,
            'total_pages': project.total_pages or 0,
            'total_pages_with_refs': project.total_pages_with_refs or 0,
            'date_created': project.created_at.strftime('%Y-%m-%d'),
            'updated_at': project.updated_at.strftime('%Y-%m-%d'),
            'owner': project.owner.username,
            # Additional fields for UI compatibility
            'folderName': project.project_folder_name,
            'path': project.project_path,
            'oversized': project.has_oversized,
            'data_dir': project.output_dir,
            'index_path': project.comlist_path
        }
        
        return JsonResponse({
            'status': 'success',
            'project': project_data
        })
        
    except Project.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'Project not found'
        }, status=404)
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)

@csrf_exempt
@login_required
def update_project(request, project_id):
    """
    API endpoint to update a project.
    
    Args:
        project_id: Project ID (URL parameter)
        Various project attributes in JSON request body
        
    Returns:
        JSON response with status and message
    """
    if request.method == 'PUT':
        try:
            project = Project.objects.get(id=project_id, owner=request.user)
            data = json.loads(request.body)
            
            # Update project fields if provided in the request
            if 'archive_id' in data:
                project.archive_id = data['archive_id']
            if 'location' in data:
                project.location = data['location']
            if 'doc_type' in data:
                project.doc_type = data['doc_type']
            if 'project_path' in data:
                project.project_path = data['project_path']
            if 'project_folder_name' in data:
                project.project_folder_name = data['project_folder_name']
            if 'comlist_path' in data:
                project.comlist_path = data['comlist_path']
            if 'output_dir' in data:
                project.output_dir = data['output_dir']
            if 'has_pdf_folder' in data:
                project.has_pdf_folder = data['has_pdf_folder']
            if 'processing_complete' in data:
                project.processing_complete = data['processing_complete']
            if 'retain_sources' in data:
                project.retain_sources = data['retain_sources']
            if 'add_to_database' in data:
                project.add_to_database = data['add_to_database']
            if 'has_oversized' in data:
                project.has_oversized = data['has_oversized']
            if 'total_pages' in data:
                project.total_pages = data['total_pages']
            if 'total_pages_with_refs' in data:
                project.total_pages_with_refs = data['total_pages_with_refs']
            if 'pdf_folder_path' in data:
                project.pdf_folder_path = data['pdf_folder_path']
                
            # Save the updated project
            project.save()
            
            return JsonResponse({
                'status': 'success',
                'message': f'Project {project.archive_id} updated successfully'
            })
            
        except Project.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'message': 'Project not found'
            }, status=404)
            
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=400)
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@csrf_exempt
@login_required
def delete_project(request, project_id):
    """
    API endpoint to delete a project.
    
    Args:
        project_id: Project ID (URL parameter)
        
    Returns:
        JSON response with status and message
    """
    if request.method == 'DELETE':
        try:
            project = Project.objects.get(id=project_id, owner=request.user)
            archive_id = project.archive_id
            project.delete()
            
            return JsonResponse({
                'status': 'success',
                'message': f'Project {archive_id} deleted successfully'
            })
            
        except Project.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'message': 'Project not found'
            }, status=404)
            
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=400)
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@login_required
def get_database_stats(request):
    """
    API endpoint to get database statistics.
    
    Returns:
        JSON response with database statistics
    """
    try:
        # Get counts for each entity
        total_projects = Project.objects.filter(owner=request.user).count()
        
        # For now, we only have Project model, but we can add more in the future
        total_rolls = 0  # Will be implemented when Roll model is created
        total_documents = 0  # Will be implemented when Document model is created
        
        return JsonResponse({
            'status': 'success',
            'stats': {
                'total_projects': total_projects,
                'total_rolls': total_rolls,
                'total_documents': total_documents
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)

@login_required
def get_document_types(request):
    """
    API endpoint to get distinct document types from the database.
    
    Returns:
        JSON response with list of document types
    """
    try:
        # Get distinct document types from projects owned by the user
        document_types = Project.objects.filter(
            owner=request.user,
            doc_type__isnull=False
        ).exclude(doc_type='').values_list('doc_type', flat=True).distinct().order_by('doc_type')
        
        # Convert QuerySet to list
        document_types_list = list(document_types)
        
        # If no document types found, return some default ones
        if not document_types_list:
            document_types_list = ["Archive", "Document", "Book", "Newspaper", "Correspondence", "Records"]
        
        return JsonResponse({
            'status': 'success',
            'document_types': document_types_list
        })
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e),
            'document_types': ["Archive", "Document", "Book", "Newspaper", "Correspondence", "Records"]
        }, status=400)

@login_required
def get_locations(request):
    """
    API endpoint to get distinct locations from the database.
    
    Returns:
        JSON response with list of locations
    """
    try:
        # Get distinct locations from projects owned by the user
        locations = Project.objects.filter(
            owner=request.user,
            location__isnull=False
        ).exclude(location='').values_list('location', flat=True).distinct().order_by('location')
        
        # Convert QuerySet to list
        locations_list = list(locations)
        
        # If no locations found, return some default ones
        if not locations_list:
            locations_list = ["OU", "DW", "Main Archive", "Satellite Office", "Remote Storage"]
        
        return JsonResponse({
            'status': 'success',
            'locations': locations_list
        })
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e),
            'locations': ["OU", "DW", "Main Archive", "Satellite Office", "Remote Storage"]
        }, status=400)

@login_required
def get_project_history(request, project_id):
    """
    API endpoint to get project history/timeline.
    
    Args:
        project_id: Project ID (URL parameter)
        
    Returns:
        JSON response with project history
    """
    try:
        project = Project.objects.get(id=project_id, owner=request.user)
        
        # For now, create a basic history based on project data
        # In the future, this could be expanded with a proper audit log/history model
        history = []
        
        # Add creation event
        history.append({
            'timestamp': project.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'action': 'Project Created',
            'description': f'Project {project.archive_id} was created'
        })
        
        # Add update event if different from creation
        if project.updated_at and project.updated_at != project.created_at:
            history.append({
                'timestamp': project.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
                'action': 'Project Updated',
                'description': f'Project {project.archive_id} was modified'
            })
        
        # Add processing milestones based on project status
        if project.has_pdf_folder:
            history.append({
                'timestamp': project.updated_at.strftime('%Y-%m-%d %H:%M:%S') if project.updated_at else project.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'action': 'PDF Folder Added',
                'description': 'PDF folder was configured for the project'
            })
        
        if project.processing_complete:
            history.append({
                'timestamp': project.updated_at.strftime('%Y-%m-%d %H:%M:%S') if project.updated_at else project.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'action': 'Processing Complete',
                'description': 'Project processing has been completed'
            })
        
        # Sort history by timestamp (newest first)
        history.sort(key=lambda x: x['timestamp'], reverse=True)
        
        return JsonResponse({
            'status': 'success',
            'history': history
        })
        
    except Project.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'Project not found'
        }, status=404)
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)

@login_required
def get_project_rolls(request, project_id):
    """
    API endpoint to get rolls associated with a project.
    
    Args:
        project_id: Project ID (URL parameter)
        
    Returns:
        JSON response with list of rolls
    """
    try:
        project = Project.objects.get(id=project_id, owner=request.user)
        
        # Import Roll model and get rolls for this project
        try:
            from ..models import Roll
            rolls = Roll.objects.filter(project=project).order_by('roll_id')
            
            rolls_data = []
            for roll in rolls:
                rolls_data.append({
                    'id': roll.id,
                    'roll_id': roll.roll_id,
                    'film_number': roll.film_number,
                    'film_type': roll.film_type,
                    'capacity': roll.capacity,
                    'pages_used': roll.pages_used,
                    'pages_remaining': roll.pages_remaining,
                    'status': roll.status,
                    'is_full': roll.is_full,
                    'is_partial': roll.is_partial,
                    'creation_date': roll.creation_date.isoformat() if roll.creation_date else None,
                    'has_split_documents': roll.has_split_documents,
                    'usable_capacity': roll.usable_capacity,
                    'film_number_source': roll.film_number_source
                })
            
            return JsonResponse({
                'status': 'success',
                'rolls': rolls_data
            })
            
        except ImportError:
            # Roll model not available, return empty list
            return JsonResponse({
                'status': 'success',
                'rolls': []
            })
        
    except Project.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'Project not found'
        }, status=404)
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)

def get_project_documents(request, project_id):
    """
    API endpoint to get documents associated with a project.
    
    Args:
        project_id: Project ID (URL parameter)
        
    Returns:
        JSON response with list of documents
    """
    try:
        project = Project.objects.get(id=project_id)
        
        # Import Document model and get documents for this project
        try:
            from ..models import Document
            documents = Document.objects.filter(project=project).order_by('doc_id')
            
            documents_data = []
            for doc in documents:
                documents_data.append({
                    'id': doc.id,
                    'doc_id': doc.doc_id,
                    'path': doc.path,
                    'com_id': doc.com_id,
                    'pages': doc.pages,
                    'has_oversized': doc.has_oversized,
                    'total_oversized': doc.total_oversized,
                    'total_references': doc.total_references,
                    'is_split': doc.is_split,
                    'roll_count': doc.roll_count,
                    'created_at': doc.created_at.isoformat() if doc.created_at else None,
                    'updated_at': doc.updated_at.isoformat() if doc.updated_at else None,
                    'name': doc.path.split('/')[-1] if doc.path else f"Document {doc.doc_id}"
                })
            
            return JsonResponse({
                'status': 'success',
                'documents': documents_data
            })
            
        except ImportError:
            # Document model not available, return empty list
            return JsonResponse({
                'status': 'success',
                'documents': []
            })
        
    except Project.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'Project not found'
        }, status=404)
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)

@login_required
@require_http_methods(["POST"])
def export_projects_excel(request):
    """
    API endpoint to export projects as Excel file.
    
    Args:
        request: HTTP request object with JSON data
        
    Returns:
        Excel file as response
    """
    try:
        import json
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        
        # Parse JSON data from request
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({
                "status": "error",
                "message": "Invalid JSON data"
            }, status=400)
        
        projects = data.get('projects', [])
        fields = data.get('fields', [])
        field_names = data.get('fieldNames', [])
        options = data.get('options', {})
        
        if not projects:
            return JsonResponse({
                "status": "error",
                "message": "No projects data provided"
            }, status=400)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Projects"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add headers
        for col_idx, field_name in enumerate(field_names, 1):
            cell = ws.cell(row=1, column=col_idx, value=field_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data rows
        for row_idx, project in enumerate(projects, 2):
            for col_idx, field in enumerate(fields, 1):
                value = project.get(field, '')
                
                # Convert boolean values to Yes/No
                if isinstance(value, bool):
                    value = 'Yes' if value else 'No'
                
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for col_idx in range(1, len(fields) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            
            for row in ws[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add auto-filter if requested
        if options.get('autoFilter', True):
            ws.auto_filter.ref = ws.dimensions
        
        # Freeze header row if requested
        if options.get('freezeHeader', True):
            ws.freeze_panes = 'A2'
        
        # Add summary statistics if requested
        if options.get('includeStats', False):
            # Add a new worksheet for statistics
            stats_ws = wb.create_sheet("Statistics")
            
            stats_data = [
                ["Statistic", "Value"],
                ["Total Projects", len(projects)],
                ["Projects with PDF Folder", sum(1 for p in projects if p.get('has_pdf_folder'))],
                ["Projects with Oversized", sum(1 for p in projects if p.get('has_oversized'))],
                ["Processing Complete", sum(1 for p in projects if p.get('processing_complete'))],
                ["Film Allocation Complete", sum(1 for p in projects if p.get('film_allocation_complete'))],
                ["Total Pages", sum(p.get('total_pages', 0) for p in projects)],
            ]
            
            for row_idx, (stat, value) in enumerate(stats_data, 1):
                stats_ws.cell(row=row_idx, column=1, value=stat)
                stats_ws.cell(row=row_idx, column=2, value=value)
                
                if row_idx == 1:  # Header row
                    stats_ws.cell(row=row_idx, column=1).font = header_font
                    stats_ws.cell(row=row_idx, column=1).fill = header_fill
                    stats_ws.cell(row=row_idx, column=2).font = header_font
                    stats_ws.cell(row=row_idx, column=2).fill = header_fill
            
            # Auto-adjust column widths for stats
            for col_idx in range(1, 3):
                column_letter = get_column_letter(col_idx)
                max_length = max(len(str(cell.value)) for cell in stats_ws[column_letter] if cell.value)
                stats_ws.column_dimensions[column_letter].width = max_length + 2
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="projects_export.xlsx"'
        
        return response
        
    except Exception as e:
        logger.error(f"Error exporting projects as Excel: {str(e)}")
        return JsonResponse({
            "status": "error",
            "message": f"Failed to export as Excel: {str(e)}"
        }, status=500)

@login_required
@require_http_methods(["POST"])
def export_projects_pdf(request):
    """
    API endpoint to export projects as PDF file.
    
    Args:
        request: HTTP request object with JSON data
        
    Returns:
        PDF file as response
    """
    try:
        import json
        from io import BytesIO
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4, legal, landscape, portrait
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from django.http import HttpResponse
        from datetime import datetime
        
        # Parse JSON data from request
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({
                "status": "error",
                "message": "Invalid JSON data"
            }, status=400)
        
        projects = data.get('projects', [])
        fields = data.get('fields', [])
        field_names = data.get('fieldNames', [])
        options = data.get('options', {})
        
        if not projects:
            return JsonResponse({
                "status": "error",
                "message": "No projects data provided"
            }, status=400)
        
        # Create PDF
        output = BytesIO()
        
        # Set page size and orientation
        page_size = letter  # Default
        if options.get('pageSize') == 'a4':
            page_size = A4
        elif options.get('pageSize') == 'legal':
            page_size = legal
        
        if options.get('orientation') == 'landscape':
            page_size = landscape(page_size)
        else:
            page_size = portrait(page_size)
        
        doc = SimpleDocTemplate(output, pagesize=page_size, topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        # Build story
        story = []
        styles = getSampleStyleSheet()
        
        # Add header if requested
        if options.get('includeHeaderFooter', True):
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            story.append(Paragraph("Project Export Report", title_style))
            story.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
            story.append(Spacer(1, 20))
        
        # Prepare table data
        table_data = [field_names]  # Header row
        
        for project in projects:
            row = []
            for field in fields:
                value = project.get(field, '')
                
                # Convert boolean values to Yes/No
                if isinstance(value, bool):
                    value = 'Yes' if value else 'No'
                
                # Truncate long text for PDF
                if isinstance(value, str) and len(value) > 50:
                    value = value[:47] + "..."
                
                row.append(str(value))
            table_data.append(row)
        
        # Create table
        table = Table(table_data)
        
        # Style the table
        table_style = TableStyle([
            # Header row styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Data rows styling
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            
            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
        ])
        
        table.setStyle(table_style)
        story.append(table)
        
        # Add footer if requested
        if options.get('includeHeaderFooter', True):
            story.append(Spacer(1, 20))
            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=8,
                alignment=1  # Center alignment
            )
            story.append(Paragraph(f"Total Projects: {len(projects)}", footer_style))
            story.append(Paragraph("Microfilm Management System", footer_style))
        
        # Build PDF
        doc.build(story)
        output.seek(0)
        
        # Create response
        response = HttpResponse(output.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="projects_export.pdf"'
        
        return response
        
    except Exception as e:
        logger.error(f"Error exporting projects as PDF: {str(e)}")
        return JsonResponse({
            "status": "error",
            "message": f"Failed to export as PDF: {str(e)}"
        }, status=500) 