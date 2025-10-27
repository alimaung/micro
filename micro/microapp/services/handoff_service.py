"""
Handoff Service for Microfilm Processing System
Handles validation, file generation, and email delivery for completed projects.
"""

import os
import re
import json
import logging
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any
from dataclasses import dataclass
from datetime import datetime
import pandas as pd
import win32com.client as win32
import uuid


# Django imports for template rendering
from django.template.loader import render_to_string
from django.template import Context, Template
from django.contrib.auth.models import User
from microapp.models import HandoffRecord

logger = logging.getLogger(__name__)

# Default email recipients
DEFAULT_EMAIL_RECIPIENTS = {
    'to': 'Dilek.kursun@rolls-royce.com', 
    'cc': 'jan.becker@rolls-royce.com; thomas.lux@rolls-royce.com',
    'bcc': 'michael.wuske@rolls-royce.com; tetiana.isakii@rolls-royce.com; shmaila.aslam@rolls-royce.com'
}

def validate_barcode(barcode: str) -> Tuple[bool, str]:
    """
    Validate that barcode is exactly 16 digits.
    
    Args:
        barcode: Barcode string to validate
        
    Returns:
        Tuple of (is_valid, normalized_barcode)
    """
    if not barcode:
        return False, ""
    
    # Remove any whitespace and convert to string
    clean_barcode = str(barcode).strip()
    
    # Check if it's exactly 16 digits
    if re.match(r'^\d{16}$', clean_barcode):
        return True, clean_barcode
    
    # Try to extract 16 consecutive digits
    digits_only = re.sub(r'\D', '', clean_barcode)
    if len(digits_only) == 16:
        return True, digits_only
    
    return False, clean_barcode

def validate_com_id(com_id: Any) -> Tuple[bool, str]:
    """
    Validate that COM ID is exactly 8 digits.
    
    Args:
        com_id: COM ID to validate (can be string, int, float)
        
    Returns:
        Tuple of (is_valid, normalized_com_id)
    """
    if com_id is None or pd.isna(com_id):
        return False, ""
    
    # Handle different input types
    if isinstance(com_id, (int, float)):
        if pd.isna(com_id) or com_id == float('inf') or com_id == float('-inf'):
            return False, ""
        # Convert to integer if it's a whole number
        if isinstance(com_id, float) and com_id.is_integer():
            com_id = int(com_id)
        clean_com_id = str(com_id)
    else:
        clean_com_id = str(com_id).strip()
    
    # Remove any non-digit characters
    digits_only = re.sub(r'\D', '', clean_com_id)
    
    # Check if it's exactly 8 digits
    if len(digits_only) == 8:
        return True, digits_only
    
    # Check if it's less than 8 digits and pad with leading zeros
    if len(digits_only) <= 8 and len(digits_only) > 0:
        padded_com_id = digits_only.zfill(8)
        return True, padded_com_id
    
    return False, clean_com_id

def remove_pdf_extension(doc_id: str) -> str:
    """
    Remove only the .pdf extension from document ID, preserving suffixes.
    
    This is used for film log lookups where documents with suffixes are filmed separately.
    For example: 1422022600000227_001.PDF -> 1422022600000227_001
    
    Args:
        doc_id: Document ID to process
        
    Returns:
        Document ID without .pdf extension but with suffix preserved
    """
    if not doc_id:
        return ""
    
    # Convert to string and strip whitespace
    normalized = str(doc_id).strip()
    
    # Remove .pdf extension (case-insensitive)
    if normalized.lower().endswith('.pdf'):
        normalized = normalized[:-4]
    
    return normalized

def normalize_document_id(doc_id: str) -> str:
    """
    Normalize document ID by removing .pdf extension and suffixes like _001, _002.
    
    This handles cases where additional documents are added after filming with suffixes.
    For example: 1422022600000227_001.PDF -> 1422022600000227
    
    This should ONLY be used for COM ID lookups, NOT for film log lookups.
    
    Args:
        doc_id: Document ID to normalize
        
    Returns:
        Normalized document ID (base barcode without suffix)
    """
    if not doc_id:
        return ""
    
    # First remove .pdf extension
    normalized = remove_pdf_extension(doc_id)
    
    # Remove suffixes like _001, _002, _003, etc. (underscore + digits)
    # This allows documents added after filming to match their base barcode in COM list
    suffix_pattern = r'_\d{3,}$'  # Match underscore followed by 3+ digits at end
    normalized = re.sub(suffix_pattern, '', normalized)
    
    return normalized

@dataclass
class ValidationResult:
    """Result of validating a single document entry."""
    document_id: str
    roll: str
    barcode: str
    com_id: str
    temp_blip: str
    film_blip: Optional[str]
    status: str  # 'validated', 'warning', 'error', 'pending'
    message: Optional[str] = None
    start_frame: Optional[int] = None
    end_frame: Optional[int] = None
    pages: Optional[int] = None

@dataclass
class ValidationSummary:
    """Summary of validation results."""
    total: int
    validated: int
    warnings: int
    errors: int
    pending: int

@dataclass
class FilmLogEntry:
    """Entry from a film log file."""
    roll_number: str
    document_id: str
    barcode: str
    blip: str
    start_frame: int
    end_frame: int
    pages: int
    timestamp: Optional[datetime] = None

class HandoffService:
    """Service for handling project handoff operations."""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def validate_project(self, project, validation_data: List[Dict]) -> Tuple[ValidationSummary, List[ValidationResult]]:
        """
        Validate project index against film log files.
        
        Args:
            project: Django Project model instance
            validation_data: List of temporary index entries to validate
            
        Returns:
            Tuple of (ValidationSummary, List[ValidationResult])
        """
        try:
            # Debug: Log incoming validation data
            self.logger.info(f"=== VALIDATION DEBUG START for project {project.archive_id} ===")
            self.logger.info(f"Received {len(validation_data)} validation entries")
            
            # Check for NaN values in incoming data
            nan_count = 0
            for i, item in enumerate(validation_data):
                com_id = item.get('com_id')
                if com_id is not None:
                    if pd.isna(com_id):
                        nan_count += 1
                        self.logger.warning(f"NaN com_id found in validation_data[{i}]: {item}")
                    elif isinstance(com_id, float) and (com_id == float('inf') or com_id == float('-inf')):
                        self.logger.warning(f"INF com_id found in validation_data[{i}]: {item}")
                else:
                    self.logger.debug(f"NULL com_id found in validation_data[{i}]: {item}")
            
            if nan_count > 0:
                self.logger.warning(f"Found {nan_count} NaN com_id values in incoming validation data")
            
            # Parse film log files
            film_logs = self._parse_film_logs(project)
            
            # Create lookup dictionary for fast access with multiple key variations
            film_log_lookup = {}
            for log_entry in film_logs:
                # Get original document ID
                doc_id = log_entry.document_id
                
                # Create primary key with original format
                key_original = f"{log_entry.roll_number}_{doc_id}"
                film_log_lookup[key_original] = log_entry
                
                # Create key with lowercase document ID
                key_lower = f"{log_entry.roll_number}_{doc_id.lower()}"
                if key_lower != key_original:
                    film_log_lookup[key_lower] = log_entry
                
                # Create key with document ID without extension (case insensitive)
                doc_id_no_ext = doc_id
                if doc_id_no_ext.lower().endswith('.pdf'):
                    doc_id_no_ext = doc_id_no_ext[:-4]
                key_no_ext = f"{log_entry.roll_number}_{doc_id_no_ext}"
                if key_no_ext != key_original and key_no_ext != key_lower:
                    film_log_lookup[key_no_ext] = log_entry
                
                # Create key with lowercase document ID without extension
                key_lower_no_ext = f"{log_entry.roll_number}_{doc_id_no_ext.lower()}"
                if key_lower_no_ext != key_original and key_lower_no_ext != key_lower and key_lower_no_ext != key_no_ext:
                    film_log_lookup[key_lower_no_ext] = log_entry
            
            # Debug lookup dictionary
            self.logger.info(f"Created lookup dictionary with {len(film_log_lookup)} keys from {len(film_logs)} log entries")
            
            # Validate each entry
            results = []
            summary = ValidationSummary(
                total=len(validation_data),
                validated=0,
                warnings=0,
                errors=0,
                pending=0
            )
            
            for i, item in enumerate(validation_data):
                # Debug: Log each item being validated
                com_id = item.get('com_id')
                com_id_type = type(com_id).__name__
                com_id_is_nan = pd.isna(com_id) if com_id is not None else False
                
                if i < 5 or com_id_is_nan:  # Log first 5 items and any with NaN
                    self.logger.debug(f"Validating item {i}: doc_id={item.get('document_id')}, "
                                    f"com_id={com_id} (type: {com_id_type}, isNaN: {com_id_is_nan})")
                
                result = self._validate_single_entry(item, film_log_lookup)
                results.append(result)
                
                # Debug: Check if result has NaN values
                if pd.isna(result.com_id) if result.com_id is not None else False:
                    self.logger.warning(f"ValidationResult {i} has NaN com_id: {result}")
                
                # Update summary
                if result.status == 'validated':
                    summary.validated += 1
                elif result.status == 'warning':
                    summary.warnings += 1
                elif result.status == 'error':
                    summary.errors += 1
                else:
                    summary.pending += 1
            
            self.logger.info(f"Validation complete for project {project.archive_id}: "
                           f"{summary.validated} validated, {summary.warnings} warnings, {summary.errors} errors")
            self.logger.info(f"=== VALIDATION DEBUG END ===")
            
            # Store validation results for handoff record creation
            validation_results_dict = {
                'total': summary.total,
                'validated': summary.validated,
                'warnings': summary.warnings,
                'errors': summary.errors,
                'pending': summary.pending
            }
            self.store_validation_results(validation_results_dict, validation_data)
            
            return summary, results
            
        except Exception as e:
            self.logger.error(f"Error validating project {project.archive_id}: {e}")
            raise
    
    def _parse_film_logs(self, project) -> List[FilmLogEntry]:
        """
        Parse film log files from project/.filmlogs/ directory.
        
        Args:
            project: Django Project model instance
            
        Returns:
            List of FilmLogEntry objects
        """
        film_logs = []
        
        if not project.project_path:
            self.logger.warning(f"No project path set for project {project.archive_id}")
            return film_logs
        
        filmlogs_dir = Path(project.project_path) / '.filmlogs'
        
        if not filmlogs_dir.exists():
            self.logger.warning(f"Film logs directory not found: {filmlogs_dir}")
            return film_logs
        
        # Find all log files (typically named like roll numbers: 10000001.txt, 10000002.txt, etc.)
        log_files = list(filmlogs_dir.glob('*.txt'))
        
        if not log_files:
            self.logger.warning(f"No log files found in {filmlogs_dir}")
            return film_logs
        
        for log_file in log_files:
            try:
                roll_number = log_file.stem  # Filename without extension
                entries = self._parse_single_log_file(log_file, roll_number)
                film_logs.extend(entries)
                self.logger.debug(f"Parsed {len(entries)} entries from {log_file}")
            except Exception as e:
                self.logger.error(f"Error parsing log file {log_file}: {e}")
                continue
        
        self.logger.info(f"Parsed {len(film_logs)} total entries from {len(log_files)} log files")
        return film_logs
    
    def _parse_single_log_file(self, log_file: Path, roll_number: str) -> List[FilmLogEntry]:
        """
        Parse a single film log file using the same logic as indexer_v2.py.
        
        Expected format:
        Rollennummer=10000001
        Datum/Zeit=5/21/2025 1:38:58 PM
        Benutzer=user1
        Computer=MICROFILM
        1427004807000278.pdf;0-1-0;0.00;1
        1427004807000278.pdf;0-1-1;0.01;2
        
        Args:
            log_file: Path to the log file
            roll_number: Roll number extracted from filename
            
        Returns:
            List of FilmLogEntry objects
        """
        entries = []
        
        try:
            # Use the same encoding detection logic as indexer_v2.py
            encoding = 'utf-16le'  # Primary encoding
            try:
                with open(log_file, 'r', encoding=encoding) as f:
                    lines = f.readlines()
            except UnicodeError:
                # Fallback to UTF-8 if UTF-16LE fails
                encoding = 'utf-8'
                with open(log_file, 'r', encoding=encoding) as f:
                    lines = f.readlines()
                    
            self.logger.debug(f"Successfully read {log_file} with encoding: {encoding}")

            # Detect if header exists by checking first few lines
            has_header = False
            header_lines = 0
            roll = date = user = ""
            
            # Check first 4 lines for header patterns
            for i, line in enumerate(lines[:4]):
                if not line.strip():
                    continue
                if "Rollennummer=" in line:
                    roll = line.split("=")[1].strip()
                    has_header = True
                elif "Datum/Zeit=" in line:
                    date = line.split("=")[1].strip()
                    has_header = True
                elif "Benutzer=" in line:
                    user = line.split("=")[1].strip()
                    has_header = True
                elif "Computer=" in line:
                    has_header = True
            
            # Determine how many lines to skip
            if has_header:
                # Count actual header lines (usually 4: Rollennummer, Datum/Zeit, Benutzer, Computer)
                header_lines = 0
                for line in lines[:6]:  # Check up to 6 lines to be safe
                    if any(keyword in line for keyword in ["Rollennummer=", "Datum/Zeit=", "Benutzer=", "Computer="]):
                        header_lines += 1
                    elif line.strip() and header_lines > 0:
                        # Found a non-header line after seeing header lines
                        break
                
                # Ensure we skip at least 4 lines if we detected a header
                header_lines = max(header_lines, 4)
                self.logger.debug(f"Header detected in {log_file}, skipping {header_lines} lines")
            else:
                header_lines = 0
                self.logger.debug(f"No header detected in {log_file}, starting from line 0")
            
            if not roll:
                self.logger.warning(f"Could not extract roll number from {log_file}")
                roll = roll_number  # Use filename as fallback
            
            # Compile regex for better performance (same as indexer_v2.py)
            pdf_pattern = re.compile(r'^\d+\.pdf;')
            
            # Process page entries (same logic as indexer_v2.py)
            page_entries = []
            
            for line in lines[header_lines:]:
                if not line.strip() or not (pdf_pattern.match(line.lower()) or '.pdf;' in line.lower()):
                    continue
                    
                parts = line.strip().split(";")
                if len(parts) < 2:
                    continue
                    
                document_filename = parts[0]  # e.g., "1427004807000278.pdf"
                blip_position = parts[1]      # e.g., "0-1-0"
                
                # Remove BOM character if present (common in UTF-16 files)
                if document_filename.startswith('\ufeff'):
                    document_filename = document_filename[1:]
                
                # Extract document ID (remove .pdf extension, case-insensitive)
                barcode = document_filename
                if barcode.lower().endswith('.pdf'):
                    barcode = barcode[:-4]  # Remove .pdf or .PDF
                
                page_entries.append({
                    'barcode': barcode,
                    'blip': blip_position,
                    'filename': document_filename,
                    'roll': roll  # Ensure roll number is associated with each entry
                })
            
            # Process blips using the same logic as indexer_v2.py
            processed_entries = self._process_blips_like_indexer(page_entries)
            
            # Create FilmLogEntry objects for each document
            for processed_entry in processed_entries:
                # Get document info
                barcode = processed_entry['barcode']
                processed_blip = processed_entry['blip']  # e.g., "0-2-21"
                document_roll = processed_entry.get('roll', roll)  # Get roll from entry or use header roll
                
                # Count pages for this document in original entries
                doc_pages = len([e for e in page_entries if e['barcode'] == barcode])
                
                # Format blip like indexer_v2.py: roll-{doc_num:04}.{page_num:05}
                blip_parts = processed_blip.split('-')
                if len(blip_parts) >= 3:
                    try:
                        doc_num = int(blip_parts[1])
                        page_num = int(blip_parts[2])
                        formatted_blip = f"{document_roll}-{doc_num:04d}.{page_num:05d}"
                    except ValueError:
                        formatted_blip = f"{document_roll}-{processed_blip}"
                else:
                    formatted_blip = f"{document_roll}-{processed_blip}"
                
                # Calculate frame range (approximate)
                start_frame = (page_num - 1) * doc_pages + 1 if page_num > 0 else 1
                end_frame = start_frame + doc_pages - 1
                
                entry = FilmLogEntry(
                    roll_number=document_roll,
                    document_id=barcode,
                    barcode=barcode,
                    blip=formatted_blip,
                    start_frame=start_frame,
                    end_frame=end_frame,
                    pages=doc_pages
                )
                entries.append(entry)
        
        except Exception as e:
            self.logger.error(f"Error reading log file {log_file}: {e}")
            raise
        
        self.logger.info(f"Parsed {len(entries)} document entries from {log_file}")
        return entries
    
    def _process_blips_like_indexer(self, page_entries: List[Dict]) -> List[Dict]:
        """
        Process blips using the same logic as indexer_v2.py process_blips method.
        
        Args:
            page_entries: List of page entry dictionaries
            
        Returns:
            List of processed entry dictionaries
        """
        from collections import defaultdict
        
        processed_entries = []
        total_pages = 0
        
        # Group by barcode (same as indexer_v2.py)
        barcode_groups = defaultdict(list)
        for entry in page_entries:
            barcode_groups[entry['barcode']].append(entry)
        
        # Process each document group with sequential document numbering
        doc_index = 1  # Start document numbering from 1
        
        for barcode, doc_entries in barcode_groups.items():
            try:
                # Get roll number from the first entry of this document
                roll = doc_entries[0].get('roll', '')
                
                # Create new blip with sequential document numbering and cumulative page numbering
                new_blip = f"0-{doc_index}-{total_pages + 1}"
                
                processed_entries.append({
                    'barcode': barcode,
                    'blip': new_blip,
                    'roll': roll  # Preserve roll number from original entry
                })
                
                total_pages += len(doc_entries)
                doc_index += 1  # Increment document index for next document
                
            except (IndexError, ValueError) as e:
                self.logger.error(f"Error processing blip for barcode {barcode}: {e}")
                continue
        
        return processed_entries
    
    def _validate_single_entry(self, temp_entry: Dict, film_log_lookup: Dict[str, FilmLogEntry]) -> ValidationResult:
        """
        Validate a single temporary index entry against film logs.
        
        Args:
            temp_entry: Temporary index entry dictionary
            film_log_lookup: Dictionary of film log entries keyed by roll_document
            
        Returns:
            ValidationResult object
        """
        document_id = temp_entry.get('document_id', '')
        roll = temp_entry.get('roll', '')
        barcode = temp_entry.get('barcode', '')
        com_id = temp_entry.get('com_id', '')
        temp_blip = temp_entry.get('temp_blip', '')
        
        # IMPORTANT: For film log lookups, we need to preserve suffixes like _001, _002
        # because these are filmed as separate documents with different blips.
        # Only use normalize_document_id() for COM ID validation (base barcode).
        
        # Remove only .pdf extension, keep suffixes for film log lookup
        doc_id_for_film_lookup = remove_pdf_extension(document_id)
        
        # For COM ID validation, normalize completely (remove suffix too)
        normalized_doc_id_for_com = normalize_document_id(document_id)
        
        # Validate barcode - normalize for validation but keep original for lookup
        barcode_for_film_lookup = remove_pdf_extension(barcode)
        normalized_barcode_for_validation = normalize_document_id(barcode)
        
        # Validate and normalize barcode (should be 16 digits after removing suffix)
        barcode_valid, normalized_barcode = validate_barcode(normalized_barcode_for_validation)
        if not barcode_valid and normalized_barcode_for_validation:
            self.logger.warning(f"Invalid barcode format: '{barcode}' (base: '{normalized_barcode_for_validation}') for document {document_id}. Expected 16 digits.")
        
        # Validate and normalize COM ID (should be 8 digits)
        com_id_valid, normalized_com_id = validate_com_id(com_id)
        if not com_id_valid and com_id:
            self.logger.warning(f"Invalid COM ID format: '{com_id}' for document {document_id}. Expected 8 digits.")
        
        # Use document ID WITH suffix for film log lookup (preserves _001, _002, etc.)
        lookup_barcode = barcode_for_film_lookup if barcode_for_film_lookup else doc_id_for_film_lookup.lower()
        
        # Create lookup key using document ID with suffix preserved
        lookup_key = f"{roll}_{lookup_barcode}"
        
        # Try alternate keys if the main one isn't found
        alternate_keys = []
        
        # Check if main key exists in lookup
        key_found = lookup_key in film_log_lookup
        
        # If not found, try some alternate variations
        if not key_found:
            # Try with original barcode if different from normalized
            if barcode and barcode != normalized_barcode:
                alt_key = f"{roll}_{barcode.lower()}"
                if alt_key != lookup_key:
                    alternate_keys.append(alt_key)
            
            # Try with normalized document ID
            if normalized_doc_id.lower() != lookup_barcode:
                alt_key2 = f"{roll}_{normalized_doc_id.lower()}"
                alternate_keys.append(alt_key2)
            
            # Try with original document ID
            alt_key3 = f"{roll}_{document_id.lower()}"
            if alt_key3 != lookup_key:
                alternate_keys.append(alt_key3)
            
            # Try with just the numeric part if it's a numeric ID
            if normalized_doc_id.isdigit():
                alt_key4 = f"{roll}_{normalized_doc_id}"
                alternate_keys.append(alt_key4)
            
            # Check all alternate keys
            for alt_key in alternate_keys:
                if alt_key in film_log_lookup:
                    lookup_key = alt_key
                    key_found = True
                    self.logger.info(f"Found document using alternate key: {alt_key}")
                    break
        
        self.logger.debug(f"Looking up key: {lookup_key} (original doc_id: {document_id}, barcode: {barcode})")
        
        # Check if entry exists in film logs
        if not key_found:
            # Get list of document IDs in film logs for this roll to help with debugging
            roll_docs = [key.split('_')[1] for key in film_log_lookup.keys() if key.startswith(f"{roll}_")]
            close_matches = self._find_close_matches(lookup_barcode, roll_docs, n=3)
            
            debug_msg = f'Document {document_id} (barcode: {barcode}) not found in film logs for roll {roll}'
            if close_matches:
                debug_msg += f". Close matches: {', '.join(close_matches)}"
            
            self.logger.warning(debug_msg)
            
            # Log available keys for debugging
            available_keys = list(film_log_lookup.keys())[:5]  # Show first 5 keys
            self.logger.debug(f"Key not found. Available keys (first 5): {available_keys}")
            
            return ValidationResult(
                document_id=document_id,
                roll=roll,
                barcode=normalized_barcode if barcode_valid else barcode,  # Use normalized barcode
                com_id=normalized_com_id if com_id_valid else com_id,      # Use normalized COM ID
                temp_blip=temp_blip,
                film_blip=None,
                status='error',
                message=f'Document {document_id} not found in film logs for roll {roll}'
            )
        
        film_entry = film_log_lookup[lookup_key]
        
        # Compare blips
        if temp_blip == film_entry.blip:
            # Perfect match
            return ValidationResult(
                document_id=document_id,
                roll=roll,
                barcode=normalized_barcode if barcode_valid else barcode,  # Use normalized barcode
                com_id=normalized_com_id if com_id_valid else com_id,      # Use normalized COM ID
                temp_blip=temp_blip,
                film_blip=film_entry.blip,
                status='validated',
                message='Blip matches exactly',
                start_frame=film_entry.start_frame,
                end_frame=film_entry.end_frame,
                pages=film_entry.pages
            )
        else:
            # Extract document and page numbers for comparison
            try:
                # Parse temp_blip and film_blip
                temp_pattern = r'^(\d+)-(\d{4})\.(\d{5})$'
                film_pattern = r'^(\d+)-(\d{4})\.(\d{5})$'
                
                temp_match = re.match(temp_pattern, temp_blip)
                film_match = re.match(film_pattern, film_entry.blip)
                
                # If both match the pattern, do structured comparison
                if temp_match and film_match:
                    temp_roll, temp_doc, temp_page = temp_match.groups()
                    film_roll, film_doc, film_page = film_match.groups()
                    
                    # If rolls match but document numbers differ by 1 and page numbers are close,
                    # consider it a minor difference (common with sequential vs. non-sequential doc numbering)
                    if temp_roll == film_roll:
                        doc_diff = abs(int(temp_doc) - int(film_doc))
                        
                        # If document numbers differ by just 1 (common when doc numbering schemes differ)
                        if doc_diff <= 1:
                                                    return ValidationResult(
                            document_id=document_id,
                            roll=roll,
                            barcode=normalized_barcode if barcode_valid else barcode,
                            com_id=normalized_com_id if com_id_valid else com_id,
                            temp_blip=temp_blip,
                            film_blip=film_entry.blip,
                            status='warning',
                            message=f'Minor blip numbering difference: expected {temp_blip}, found {film_entry.blip}',
                            start_frame=film_entry.start_frame,
                            end_frame=film_entry.end_frame,
                            pages=film_entry.pages
                        )
            except Exception as e:
                self.logger.debug(f"Error parsing blips for structured comparison: {e}")
                # Continue with regular similarity check
            
            # Check if it's a minor difference (warning) or major difference (error)
            similarity = self._calculate_blip_similarity(temp_blip, film_entry.blip)
            
            if similarity > 0.8:  # Minor difference - warning
                return ValidationResult(
                    document_id=document_id,
                    roll=roll,
                    barcode=normalized_barcode if barcode_valid else barcode,
                    com_id=normalized_com_id if com_id_valid else com_id,
                    temp_blip=temp_blip,
                    film_blip=film_entry.blip,
                    status='warning',
                    message=f'Blip mismatch: expected {temp_blip}, found {film_entry.blip}',
                    start_frame=film_entry.start_frame,
                    end_frame=film_entry.end_frame,
                    pages=film_entry.pages
                )
            else:  # Major difference - error
                return ValidationResult(
                    document_id=document_id,
                    roll=roll,
                    barcode=normalized_barcode if barcode_valid else barcode,
                    com_id=normalized_com_id if com_id_valid else com_id,
                    temp_blip=temp_blip,
                    film_blip=film_entry.blip,
                    status='error',
                    message=f'Significant blip mismatch: expected {temp_blip}, found {film_entry.blip}',
                    start_frame=film_entry.start_frame,
                    end_frame=film_entry.end_frame,
                    pages=film_entry.pages
                )
    
    def _find_close_matches(self, doc_id: str, candidates: list, n: int = 3) -> list:
        """Find close string matches in a list of candidates."""
        try:
            import difflib
            return difflib.get_close_matches(doc_id, candidates, n=n, cutoff=0.7)
        except Exception as e:
            self.logger.debug(f"Error finding close matches: {e}")
            return []
    
    def _calculate_blip_similarity(self, blip1: str, blip2: str) -> float:
        """
        Calculate similarity between two blip strings.
        
        Args:
            blip1: First blip string
            blip2: Second blip string
            
        Returns:
            Similarity score between 0 and 1
        """
        if blip1 == blip2:
            return 1.0
        
        # Simple similarity based on common characters and structure
        if len(blip1) == 0 or len(blip2) == 0:
            return 0.0
        
        # Check if they have the same structure (roll-doc.frame format)
        pattern = r'^(\d+)-(\d{4})\.(\d{5})$'
        match1 = re.match(pattern, blip1)
        match2 = re.match(pattern, blip2)
        
        if match1 and match2:
            # Compare components
            roll1, doc1, frame1 = match1.groups()
            roll2, doc2, frame2 = match2.groups()
            
            # Same roll and document, different frame numbers = high similarity
            if roll1 == roll2 and doc1 == doc2:
                frame_diff = abs(int(frame1) - int(frame2))
                if frame_diff <= 5:  # Within 5 frames
                    return 0.9
                elif frame_diff <= 10:  # Within 10 frames
                    return 0.8
                else:
                    return 0.5
            
            # Same roll, different document, similar frame = medium similarity
            if roll1 == roll2:
                doc_diff = abs(int(doc1) - int(doc2))
                if doc_diff == 1:  # Adjacent document numbers
                    return 0.7
                elif doc_diff <= 2:  # Nearby document numbers
                    return 0.6
                else:
                    return 0.4
        
        # Fallback: simple string similarity
        common_chars = sum(1 for a, b in zip(blip1, blip2) if a == b)
        max_length = max(len(blip1), len(blip2))
        return common_chars / max_length if max_length > 0 else 0.0
    
    def generate_handoff_files(self, project, validated_results: List[ValidationResult]) -> Dict[str, Any]:
        """
        Generate the final handoff files (Excel only - DAT file generation disabled).
        
        Args:
            project: Django Project model instance
            validated_results: List of ValidationResult objects
            
        Returns:
            Dictionary with file paths and metadata
        """
        try:
            # Create export directory
            project_path = Path(project.project_path) if project.project_path else Path.cwd()
            export_dir = project_path / '.export'
            export_dir.mkdir(exist_ok=True)
            
            # Filter successful validations (validated + warnings)
            valid_entries = [r for r in validated_results if r.status in ['validated', 'warning']]
            
            if not valid_entries:
                raise ValueError("No valid entries to export")
            
            # Generate Excel file only
            excel_path = self._generate_excel_file(export_dir, valid_entries, project)
            # DAT file generation commented out
            # dat_path = self._generate_dat_file(export_dir, valid_entries, project)
            
            # Get file sizes
            excel_size = excel_path.stat().st_size if excel_path.exists() else 0
            # dat_size = dat_path.stat().st_size if dat_path.exists() else 0
            
            self.logger.info(f"Generated handoff files for project {project.archive_id}: "
                           f"Excel ({excel_size} bytes)")
            
            return {
                'excel_path': str(excel_path),
                # 'dat_path': str(dat_path),
                'excel_size': excel_size,
                # 'dat_size': dat_size,
                'entries_count': len(valid_entries),
                'generated_at': datetime.now().isoformat()
            }
            
        except Exception as e:
            self.logger.error(f"Error generating handoff files for project {project.archive_id}: {e}")
            raise
    
    def _generate_excel_file(self, export_dir: Path, entries: List[ValidationResult], project) -> Path:
        """
        Generate Excel file with validation results.
        
        Args:
            export_dir: Directory to save the file
            entries: List of ValidationResult objects
            project: Django Project model instance
            
        Returns:
            Path to the generated Excel file
        """
        # Generate timestamp for filename
        timestamp = datetime.now().strftime('%d%m%y%H%M')
        archive_id = project.archive_id or 'PROJECT'
        excel_path = export_dir / f'{archive_id}_blips_{timestamp}.xlsx'
        
        self.logger.info(f"=== EXCEL GENERATION DEBUG START for {archive_id} ===")
        self.logger.info(f"Processing {len(entries)} validation results for Excel export")
        
        # Prepare data for DataFrame with only required columns
        data = []
        nan_entries = []
        
        for i, entry in enumerate(entries):
            # Debug: Check for NaN values in ValidationResult
            com_id_is_nan = pd.isna(entry.com_id) if entry.com_id is not None else False
            if com_id_is_nan:
                nan_entries.append(i)
                self.logger.warning(f"NaN com_id in ValidationResult[{i}]: "
                                  f"doc_id={entry.document_id}, com_id={entry.com_id}, "
                                  f"barcode={entry.barcode}, status={entry.status}")
            
            # Normalize and validate barcode and COM ID using robust utilities
            barcode_valid, clean_barcode = validate_barcode(entry.barcode)
            if not barcode_valid:
                # Fallback: normalize document ID if barcode is invalid
                clean_barcode = normalize_document_id(entry.barcode or '')
                if not clean_barcode:
                    clean_barcode = str(entry.barcode) if entry.barcode is not None else ''
                self.logger.warning(f"Invalid barcode format for entry {i}: '{entry.barcode}'. Expected 16 digits.")
            
            # Validate and normalize COM ID (should be 8 digits)
            com_id_valid, clean_com_id = validate_com_id(entry.com_id)
            if not com_id_valid and entry.com_id:
                self.logger.warning(f"Invalid COM ID format for entry {i}: '{entry.com_id}'. Expected 8 digits.")
            
            # Use film_blip (actual filmed blip) if available, otherwise temp_blip
            # For warnings, we MUST use film_blip since that's the actual blip from filming
            clean_blip = entry.film_blip or entry.temp_blip or ''
            
            # Log which blip source is being used for warnings
            if entry.status == 'warning':
                if entry.film_blip:
                    self.logger.info(f"Entry {i} (warning): Using film_blip: {entry.film_blip} (temp was: {entry.temp_blip})")
                else:
                    self.logger.warning(f"Entry {i} (warning): No film_blip available, using temp_blip: {entry.temp_blip}")
            
            data.append({
                'Barcode': clean_barcode,
                'Bildnummer': clean_com_id,
                'Blipindex': clean_blip
            })
        
        if nan_entries:
            self.logger.warning(f"Found NaN com_id values in {len(nan_entries)} entries: {nan_entries}")
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Debug: Check DataFrame for NaN values
        nan_count_before = df.isna().sum().sum()
        if nan_count_before > 0:
            self.logger.warning(f"DataFrame contains {nan_count_before} NaN values before cleaning")
            for col in df.columns:
                col_nan_count = df[col].isna().sum()
                if col_nan_count > 0:
                    self.logger.warning(f"Column '{col}' has {col_nan_count} NaN values")
        
        # Additional data cleaning to ensure no NaN/INF values
        df = df.fillna('')  # Replace NaN with empty strings
        df = df.replace([float('inf'), float('-inf')], '')  # Replace INF with empty strings
        
        # Debug: Check DataFrame after cleaning
        nan_count_after = df.isna().sum().sum()
        if nan_count_after > 0:
            self.logger.warning(f"DataFrame still contains {nan_count_after} NaN values after cleaning!")
        else:
            self.logger.info(f"DataFrame successfully cleaned of NaN values")
        
        # Sort numerically by barcode (convert to int for proper numeric sorting)
        try:
            # Create a numeric version for sorting, handling non-numeric barcodes
            def safe_numeric_convert(x):
                try:
                    # Extract numeric part if barcode contains non-numeric characters
                    numeric_part = ''.join(filter(str.isdigit, str(x)))
                    return int(numeric_part) if numeric_part else 0
                except (ValueError, TypeError):
                    return 0
            
            df['barcode_numeric'] = df['Barcode'].apply(safe_numeric_convert)
            df = df.sort_values('barcode_numeric').drop('barcode_numeric', axis=1)
        except Exception as e:
            self.logger.warning(f"Error in numeric sorting, falling back to string sort: {e}")
            # Fallback to string sorting if numeric conversion fails
            df = df.sort_values('Barcode')
        
        # Save to Excel with specified headers only
        try:
            # Create Excel writer with proper options for NaN handling
            excel_options = {
                'nan_inf_to_errors': True,
                'strings_to_numbers': False,
                'strings_to_formulas': False
            }
            
            with pd.ExcelWriter(excel_path, engine='xlsxwriter', options=excel_options) as writer:
                # Write main data sheet with only the required columns
                df.to_excel(writer, sheet_name='Index', index=False)
                
                # Get workbook and worksheet objects for formatting
                workbook = writer.book
                worksheet = writer.sheets['Index']
                
                # Add header formatting - nice blue background, white font, centered, 30 row height
                header_format = workbook.add_format({
                    'bold': True,
                    'font_color': 'white',
                    'bg_color': '#10069F',  # Nice blue color
                    'align': 'center',
                    'valign': 'vcenter',
                    'text_wrap': True,
                    'border': 1,
                    'border_color': '#FFFFFF',
                    'font_name': 'Arial',
                    'font_size': 11
                })
                
                # Add data formatting - left aligned
                data_format = workbook.add_format({
                    'align': 'left',
                    'valign': 'vcenter',
                    'border': 1,
                    'border_color': '#D3D3D3',
                    'font_name': 'Arial',
                    'font_size': 10
                })
                
                # Set header row height to 30
                worksheet.set_row(0, 30)
                
                # Apply header formatting to all header cells
                for col_num, value in enumerate(df.columns.values):
                    worksheet.write(0, col_num, value, header_format)
                
                # Apply data formatting to all data rows
                for row_num in range(1, len(df) + 1):
                    for col_num in range(len(df.columns)):
                        cell_value = df.iloc[row_num - 1, col_num]
                        
                        # Ensure cell value is clean and properly formatted
                        if pd.isna(cell_value) or cell_value in [float('inf'), float('-inf')]:
                            cell_value = ''
                            self.logger.debug(f"Cleaned NaN/INF cell value at row {row_num}, col {col_num}")
                        
                        # Convert to string to ensure consistent formatting
                        if isinstance(cell_value, (int, float)) and not pd.isna(cell_value):
                            cell_value = str(int(cell_value)) if isinstance(cell_value, float) and cell_value.is_integer() else str(cell_value)
                        
                        worksheet.write(row_num, col_num, cell_value, data_format)
                
                # Auto-adjust column widths with better sizing
                for i, col in enumerate(df.columns):
                    # Calculate max width needed
                    max_length = max(
                        df[col].astype(str).map(len).max(),  # Max data length
                        len(col)  # Header length
                    )
                    # Set column width with padding, but cap at reasonable maximum
                    adjusted_width = min(max_length + 3, 50)
                    worksheet.set_column(i, i, adjusted_width)
                
                # Add some additional formatting touches
                # Freeze the header row
                worksheet.freeze_panes(1, 0)
                
                # Set print options
                worksheet.set_landscape()
                worksheet.set_paper(9)  # A4 paper
                worksheet.fit_to_pages(1, 0)  # Fit to 1 page wide, any number of pages tall
                
                # Add print margins
                worksheet.set_margins(0.7, 0.7, 0.75, 0.75)  # left, right, top, bottom
            
            self.logger.info(f"Excel file successfully created with formatting: {excel_path}")
            self.logger.info(f"=== EXCEL GENERATION DEBUG END ===")
        
        except Exception as e:
            self.logger.error(f"Error writing Excel file with xlsxwriter: {e}")
            # Fallback: try with openpyxl engine but with basic formatting
            try:
                self.logger.info("Attempting fallback with openpyxl engine")
                
                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Index', index=False)
                    
                    # Try to add basic formatting with openpyxl
                    try:
                        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
                        
                        worksheet = writer.sheets['Index']
                        
                        # Header formatting
                        header_font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
                        header_fill = PatternFill(start_color='10069F', end_color='10069F', fill_type='solid')
                        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        border = Border(
                            left=Side(style='thin', color='FFFFFF'),
                            right=Side(style='thin', color='FFFFFF'),
                            top=Side(style='thin', color='FFFFFF'),
                            bottom=Side(style='thin', color='FFFFFF')
                        )
                        
                        # Apply header formatting
                        for col in range(1, len(df.columns) + 1):
                            cell = worksheet.cell(row=1, column=col)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                            cell.border = border
                        
                        # Set header row height
                        worksheet.row_dimensions[1].height = 30
                        
                        # Data formatting
                        data_font = Font(name='Arial', size=10)
                        data_alignment = Alignment(horizontal='left', vertical='center')
                        data_border = Border(
                            left=Side(style='thin', color='D3D3D3'),
                            right=Side(style='thin', color='D3D3D3'),
                            top=Side(style='thin', color='D3D3D3'),
                            bottom=Side(style='thin', color='D3D3D3')
                        )
                        
                        # Apply data formatting
                        for row in range(2, len(df) + 2):
                            for col in range(1, len(df.columns) + 1):
                                cell = worksheet.cell(row=row, column=col)
                                cell.font = data_font
                                cell.alignment = data_alignment
                                cell.border = data_border
                        
                        # Auto-adjust column widths
                        for i, col in enumerate(df.columns, 1):
                            max_length = max(df[col].astype(str).map(len).max(), len(col))
                            worksheet.column_dimensions[worksheet.cell(row=1, column=i).column_letter].width = min(max_length + 3, 50)
                        
                    except ImportError:
                        self.logger.warning("openpyxl styling not available, creating basic Excel file")
                    except Exception as style_error:
                        self.logger.warning(f"Could not apply openpyxl styling: {style_error}")
                
                self.logger.info("Successfully created Excel file with openpyxl fallback")
            except Exception as fallback_error:
                self.logger.error(f"Fallback Excel creation also failed: {fallback_error}")
                raise
        
        return excel_path
    
    # DAT file generation disabled - function commented out
    def _generate_dat_file_DISABLED(self, export_dir: Path, entries: List[ValidationResult], project) -> Path:
        """
        Generate fixed-width DAT file for legacy systems.
        
        Args:
            export_dir: Directory to save the file
            entries: List of ValidationResult objects
            project: Django Project model instance
            
        Returns:
            Path to the generated DAT file
        """
        # Generate timestamp for filename
        timestamp = datetime.now().strftime('%d%m%y%H%M')
        archive_id = project.archive_id or 'PROJECT'
        dat_path = export_dir / f'{archive_id}_scan_{timestamp}.dat'
        
        # Prepare entries for sorting with robust validation
        entries_for_sorting = []
        for entry in entries:
            # Validate and normalize barcode (should be 16 digits)
            barcode_valid, clean_barcode = validate_barcode(entry.barcode)
            if not barcode_valid:
                # Fallback: normalize document ID if barcode is invalid
                clean_barcode = normalize_document_id(entry.barcode or '')
                if not clean_barcode:
                    clean_barcode = str(entry.barcode) if entry.barcode is not None else ''
            
            # Validate and normalize COM ID (should be 8 digits)
            com_id_valid, clean_com_id = validate_com_id(entry.com_id)
            if not com_id_valid and entry.com_id:
                clean_com_id = ''  # Use empty string for invalid COM IDs in DAT file
            
            clean_blip = entry.film_blip or entry.temp_blip or ''
            
            entries_for_sorting.append({
                'barcode': clean_barcode,
                'com_id': clean_com_id,
                'blip': clean_blip
            })
        
        # Sort numerically by barcode with improved error handling
        try:
            def safe_barcode_sort_key(entry):
                try:
                    barcode = entry['barcode']
                    # Extract numeric part if barcode contains non-numeric characters
                    numeric_part = ''.join(filter(str.isdigit, str(barcode)))
                    return int(numeric_part) if numeric_part else 0
                except (ValueError, TypeError):
                    return 0
            
            sorted_entries = sorted(entries_for_sorting, key=safe_barcode_sort_key)
        except Exception as e:
            self.logger.warning(f"Error in numeric sorting for DAT file, falling back to string sort: {e}")
            # Fallback to string sorting if numeric conversion fails
            sorted_entries = sorted(entries_for_sorting, key=lambda x: str(x['barcode']))
        
        with open(dat_path, 'w', encoding='utf-8') as f:
            # Write only data lines in fixed-width format (no header comments)
            for entry in sorted_entries:
                # Format: COM_ID (13 chars) + BARCODE (48 chars) + BLIP
                # Ensure com_id and barcode are strings before calling ljust
                com_id = str(entry['com_id'] or '').ljust(13)[:13]
                barcode = str(entry['barcode'] or '').ljust(48)[:48]
                blip = str(entry['blip'] or '')
                
                line = f"{com_id}{barcode}{blip}\n"
                f.write(line)
        
        return dat_path
    
    def send_handoff_email(self, project, email_data: Dict[str, Any], file_paths: Dict[str, Any], user: User = None) -> Dict[str, Any]:
        """
        Send handoff email using form data to populate signature placeholders.
        
        Args:
            project: Django Project model instance
            email_data: Dictionary with form data (to, cc, bcc, subject, archive_id, film_numbers, custom_message, use_form_data)
            file_paths: Dictionary with file paths from generate_handoff_files
            user: User who initiated the handoff
            
        Returns:
            Dictionary with send result
        """
        try:
            # Create handoff record to track this email
            handoff_record = self._create_handoff_record(project, email_data, file_paths, user)
            
            # Handle default recipients
            if not email_data.get('to'):
                email_data['to'] = DEFAULT_EMAIL_RECIPIENTS['to']
            
            if not email_data.get('cc'):
                email_data['cc'] = DEFAULT_EMAIL_RECIPIENTS['cc']
            
            if not email_data.get('bcc'):
                email_data['bcc'] = DEFAULT_EMAIL_RECIPIENTS['bcc']

            # Validate email addresses before proceeding
            def validate_email_list(email_string):
                if not email_string:
                    return []
                # Handle both comma and semicolon separators (convert commas to semicolons for Outlook)
                # First normalize by replacing semicolons with commas for consistent processing
                normalized_string = email_string.replace(';', ',')
                emails = [email.strip() for email in normalized_string.split(',') if email.strip()]
                valid_emails = []
                for email in emails:
                    if '@' in email and '.' in email.split('@')[1]:
                        valid_emails.append(email)
                    else:
                        self.logger.warning(f"Invalid email address format: {email}")
                return valid_emails
            
            to_emails = validate_email_list(email_data['to'])
            cc_emails = validate_email_list(email_data.get('cc', ''))
            bcc_emails = validate_email_list(email_data.get('bcc', ''))
            
            # Log email processing for debugging
            self.logger.info(f"Processing emails - TO: {to_emails}, CC: {cc_emails}, BCC: {bcc_emails}")
            
            if not to_emails:
                raise ValueError("No valid TO email addresses provided")
            
            # Update email_data with validated emails using semicolons for Outlook
            email_data['to'] = '; '.join(to_emails)
            if cc_emails:
                email_data['cc'] = '; '.join(cc_emails)
                self.logger.info(f"CC emails processed: {len(cc_emails)} recipients - {email_data['cc']}")
            else:
                email_data['cc'] = ''
                self.logger.info("No CC emails provided")
            
            if bcc_emails:
                email_data['bcc'] = '; '.join(bcc_emails)
                self.logger.info(f"BCC emails processed: {len(bcc_emails)} recipients - {email_data['bcc']}")
            else:
                email_data['bcc'] = ''
                self.logger.info("No BCC emails provided")

            # Initialize Outlook
            try:
                import pythoncom
                # Initialize COM for this thread
                pythoncom.CoInitialize()
                self.logger.debug("COM initialized successfully")
            except Exception as com_init_error:
                self.logger.warning(f"COM initialization warning: {com_init_error}")
                # Continue anyway as COM might already be initialized
            
            outlook = None
            mail = None
            msg_path = None
            msg_filename = None
            
            try:
                outlook = win32.Dispatch('Outlook.Application')
                mail = outlook.CreateItem(0)  # 0 = Mail item
                
                # Set basic email properties
                mail.To = email_data['to']
                if email_data.get('cc'):
                    mail.CC = email_data['cc']
                if email_data.get('bcc'):
                    mail.BCC = email_data['bcc']
                mail.Subject = email_data['subject']
                
                # Try to resolve recipients before proceeding
                try:
                    # This will attempt to resolve all recipients
                    mail.Recipients.ResolveAll()
                    self.logger.info("All recipients resolved successfully")
                except Exception as resolve_error:
                    self.logger.warning(f"Some recipients could not be resolved: {resolve_error}")
                    # Continue anyway - Outlook might still send the email
                    
                    # Alternative approach: Add recipients manually and mark as resolved
                    try:
                        mail.Recipients.RemoveAll()  # Clear existing recipients
                        
                        # Add TO recipients
                        to_emails = [email.strip() for email in email_data['to'].replace(';', ',').split(',') if email.strip()]
                        for email_addr in to_emails:
                            recipient = mail.Recipients.Add(email_addr)
                            recipient.Type = 1  # olTo
                            recipient.Resolve()  # Try to resolve individual recipient
                            self.logger.debug(f"Added TO recipient: {email_addr}")
                        
                        # Add CC recipients
                        if email_data.get('cc'):
                            cc_emails = [email.strip() for email in email_data['cc'].replace(';', ',').split(',') if email.strip()]
                            self.logger.info(f"Adding {len(cc_emails)} CC recipients: {cc_emails}")
                            for email_addr in cc_emails:
                                recipient = mail.Recipients.Add(email_addr)
                                recipient.Type = 2  # olCC
                                recipient.Resolve()  # Try to resolve individual recipient
                                self.logger.debug(f"Added CC recipient: {email_addr}")
                        
                        # Add BCC recipients
                        if email_data.get('bcc'):
                            bcc_emails = [email.strip() for email in email_data['bcc'].replace(';', ',').split(',') if email.strip()]
                            self.logger.info(f"Adding {len(bcc_emails)} BCC recipients: {bcc_emails}")
                            for email_addr in bcc_emails:
                                recipient = mail.Recipients.Add(email_addr)
                                recipient.Type = 3  # olBCC
                                recipient.Resolve()  # Try to resolve individual recipient
                                self.logger.debug(f"Added BCC recipient: {email_addr}")
                        
                        self.logger.info("Recipients added manually")
                    except Exception as manual_error:
                        self.logger.warning(f"Manual recipient resolution also failed: {manual_error}")
                        # Continue with original recipients
                        mail.To = email_data['to']
                        if email_data.get('cc'):
                            mail.CC = email_data['cc']
                        if email_data.get('bcc'):
                            mail.BCC = email_data['bcc']
                
                # Get signature with placeholders using GetInspector trick
                mail.GetInspector  # This triggers Outlook to add the default signature
                
                # Get the signature HTML (now contains default signature)
                signature_html = mail.HTMLBody
                
                # Prepare placeholder values
                archive_id = email_data.get('archive_id') or project.archive_id
                film_numbers = email_data.get('film_numbers', '')
                
                # If film numbers are empty, get them from the project
                if not film_numbers or film_numbers.strip() == '':
                    film_numbers = self._get_project_roll_numbers(project)
                
                custom_message = email_data.get('custom_message', '')
                
                # Generate timestamp for filenames
                timestamp = datetime.now().strftime('%d%m%y%H%M')
                
                # Replace signature placeholders
                if signature_html:
                    # Replace XXX with archive_id (appears in multiple places)
                    signature_html = signature_html.replace('XXX', archive_id)
                    
                    # Replace YYY with film numbers
                    signature_html = signature_html.replace('YYY', film_numbers)
                    
                    # Replace DDMMYYHHMM with actual timestamp
                    signature_html = signature_html.replace('DDMMYYHHMM', timestamp)
                    
                    # Replace QQQ with empty string (remove quote placeholder)
                    signature_html = signature_html.replace('QQQ', '')
                    
                    # Replace CCC with custom message or em dash if empty
                    if custom_message and custom_message.strip():
                        # Convert line breaks to HTML breaks for proper display
                        formatted_message = custom_message.strip().replace('\n', '<br>')
                        signature_html = signature_html.replace('CCC', formatted_message)
                    else:
                        # Replace CCC placeholder with em dash if no custom message
                        signature_html = signature_html.replace('CCC', '—')
                    
                    self.logger.debug(f"Replaced signature placeholders: XXX→{archive_id}, YYY→{film_numbers}, DDMMYYHHMM→{timestamp}, QQQ→(removed), CCC→{custom_message[:30] if custom_message else 'em dash'}...")
                    
                    mail.HTMLBody = signature_html
                
                # Add attachments if available (Excel only - DAT file disabled)
                if file_paths:
                    # Handle both old and new file_paths structure
                    if isinstance(file_paths, dict):
                        # New structure from generate_handoff_files - Excel only
                        if 'excel_path' in file_paths and os.path.exists(file_paths['excel_path']):
                            mail.Attachments.Add(file_paths['excel_path'])
                            self.logger.info(f"Added Excel attachment: {file_paths['excel_path']}")
                        
                        # DAT file attachment disabled
                        # if 'dat_path' in file_paths and os.path.exists(file_paths['dat_path']):
                        #     mail.Attachments.Add(file_paths['dat_path'])
                        #     self.logger.info(f"Added DAT attachment: {file_paths['dat_path']}")
                        
                        # Only process legacy structure if new structure keys are not present
                        elif not ('excel_path' in file_paths):  # Removed 'dat_path' check since it's disabled
                            # Legacy structure support - Excel files only
                            for file_type, file_info in file_paths.items():
                                if isinstance(file_info, dict) and 'path' in file_info:
                                    file_path = file_info['path']
                                    # Only attach Excel files, skip DAT files
                                    if os.path.exists(file_path) and not file_path.lower().endswith('.dat'):
                                        mail.Attachments.Add(file_path)
                                        self.logger.info(f"Added legacy attachment: {file_path}")
                                elif isinstance(file_info, str) and os.path.exists(file_info) and not file_info.lower().endswith('.dat'):
                                    mail.Attachments.Add(file_info)
                                    self.logger.info(f"Added string path attachment: {file_info}")
                
                # Debug mail object properties before sending
                try:
                    # Log mail properties to help diagnose the issue
                    self.logger.info(f"=== MAIL OBJECT DIAGNOSTIC INFO (GREEN) ===")
                    self.logger.info(f"To: {mail.To}")
                    self.logger.info(f"CC: {mail.CC}")
                    self.logger.info(f"BCC: {mail.BCC}")
                    self.logger.info(f"Subject: {mail.Subject}")
                    self.logger.info(f"Attachment count: {mail.Attachments.Count}")
                    
                    # Get list of attachments
                    attachments_list = []
                    for i in range(1, mail.Attachments.Count + 1):
                        att = mail.Attachments.Item(i)
                        attachments_list.append(f"{att.DisplayName} ({att.Type})")
                    self.logger.info(f"Attachments: {attachments_list}")
                    
                    # Check recipients
                    recipients_count = mail.Recipients.Count
                    self.logger.info(f"Recipients count: {recipients_count}")
                    resolved_count = 0
                    unresolved_count = 0
                    for i in range(1, recipients_count + 1):
                        recipient = mail.Recipients.Item(i)
                        status = "Resolved" if recipient.Resolved else "Unresolved"
                        if recipient.Resolved:
                            resolved_count += 1
                        else:
                            unresolved_count += 1
                    self.logger.info(f"Recipients resolved: {resolved_count}, unresolved: {unresolved_count}")
                    
                    # Log mail delivery options
                    self.logger.info(f"Importance: {mail.Importance}")
                    self.logger.info(f"Sensitivity: {mail.Sensitivity}")
                    
                    # Try to get message size
                    try:
                        html_length = len(mail.HTMLBody) if mail.HTMLBody else 0
                        text_length = len(mail.Body) if mail.Body else 0
                        self.logger.info(f"HTML body length: {html_length}, Text body length: {text_length}")
                    except Exception as size_error:
                        self.logger.info(f"Could not determine message size: {size_error}")
                    
                    self.logger.info(f"=== END MAIL DIAGNOSTIC INFO ===")
                except Exception as diag_error:
                    self.logger.warning(f"Error collecting mail diagnostics: {diag_error}")
                
                # Check action parameter to determine whether to send or save
                action = email_data.get('action', 'send')
                
                if action == 'save':
                    # Save the email as MSG file
                    try:
                        self.logger.info(f"Attempting to save email as MSG file...")
                        
                        # Generate timestamp for filename
                        timestamp = datetime.now().strftime('%d%m%y%H%M')
                        archive_id = email_data.get('archive_id') or project.archive_id
                        
                        # Create export directory if it doesn't exist
                        project_path = Path(project.project_path) if project.project_path else Path.cwd()
                        export_dir = project_path / '.export'
                        export_dir.mkdir(exist_ok=True)
                        
                        # Define MSG file path
                        msg_filename = f'{archive_id}_handoff_{timestamp}.msg'
                        msg_path = export_dir / msg_filename
                        
                        # Save as MSG file (3 = olMSG format)
                        mail.SaveAs(str(msg_path), 3)
                        
                        self.logger.info(f"Email saved as MSG file: {msg_path}")
                        send_method = "saved_msg"
                    except Exception as save_error:
                        self.logger.warning(f"Failed to save email as MSG file: {save_error}")
                        # Log the error details
                        if hasattr(save_error, 'excepinfo'):
                            self.logger.warning(f"Exception info: {save_error.excepinfo}")
                        
                        # Fallback: Display the email for manual saving
                        try:
                            self.logger.info(f"Falling back to displaying email for manual saving...")
                            mail.Display()
                            self.logger.info(f"Email displayed for manual saving for project {project.archive_id}")
                            send_method = "displayed"
                        except Exception as display_error:
                            self.logger.error(f"Failed to display email: {display_error}")
                            raise Exception(f"Could not save or display email: {save_error}")
                else:
                    # Send the email
                    try:
                        self.logger.info(f"Attempting to send email directly...")
                        mail.Display()
                        mail.Send()
                        self.logger.info(f"Email sent successfully for project {project.archive_id}")
                        send_method = "sent"
                    except Exception as send_error:
                        self.logger.warning(f"Failed to send email directly: {send_error}")
                        # Log the error details
                        if hasattr(send_error, 'excepinfo'):
                            self.logger.warning(f"Exception info: {send_error.excepinfo}")
                        
                        # Fallback: Display the email for manual sending
                        try:
                            self.logger.info(f"Falling back to displaying email for manual sending...")
                            mail.Display()
                            self.logger.info(f"Email displayed for manual sending for project {project.archive_id}")
                            send_method = "displayed"
                        except Exception as display_error:
                            self.logger.error(f"Failed to display email: {display_error}")
                            raise Exception(f"Could not send or display email: {send_error}")

                # Mark handoff as sent
                if send_method == "sent":
                    handoff_record.mark_sent()
                elif send_method == "saved_msg":
                    # For MSG files, mark as sent since file is ready for use
                    handoff_record.mark_sent()
                else:
                    # For displayed emails, we'll mark as sent since user will likely send it
                    handoff_record.mark_sent()
                
                # Mark project as handed off when email is successfully sent/displayed
                try:
                    project.handoff_complete = True
                    project.save(update_fields=['handoff_complete'])
                    self.logger.info(f"Marked project {project.archive_id} as handed off (handoff_complete=True)")
                except Exception as project_update_error:
                    self.logger.error(f"Failed to update project handoff status: {project_update_error}")
                    # Don't fail the entire handoff if project update fails
                
                # Prepare return message based on method
                if send_method == "saved_msg":
                    message = f'Email saved as MSG file: {msg_filename}'
                    return_data = {
                        'success': True,
                        'message': message,
                        'method': send_method,
                        'archive_id': archive_id,
                        'film_numbers': film_numbers,
                        'handoff_id': handoff_record.handoff_id,
                        'msg_file_path': str(msg_path),
                        'msg_filename': msg_filename
                    }
                elif send_method == "sent":
                    message = f'Email sent automatically to {email_data["to"]}'
                    return_data = {
                        'success': True,
                        'message': message,
                        'method': send_method,
                        'archive_id': archive_id,
                        'film_numbers': film_numbers,
                        'handoff_id': handoff_record.handoff_id
                    }
                else:
                    message = f'Email opened in Outlook for manual sending to {email_data["to"]}'
                    return_data = {
                        'success': True,
                        'message': message,
                        'method': send_method,
                        'archive_id': archive_id,
                        'film_numbers': film_numbers,
                        'handoff_id': handoff_record.handoff_id
                    }
                
                return return_data
                
            except Exception as outlook_error:
                self.logger.error(f"Outlook COM error: {outlook_error}")
                raise
            finally:
                # Clean up COM
                try:
                    import pythoncom
                    pythoncom.CoUninitialize()
                    self.logger.debug("COM uninitialized successfully")
                except Exception as cleanup_error:
                    self.logger.warning(f"COM cleanup warning: {cleanup_error}")
            
        except Exception as e:
            error_msg = f"Failed to send email for project {project.archive_id}: {str(e)}"
            self.logger.error(error_msg, exc_info=True)
            
            # Mark handoff as failed if record was created
            try:
                if 'handoff_record' in locals():
                    handoff_record.mark_failed(error_msg)
            except Exception as record_error:
                self.logger.error(f"Failed to update handoff record: {record_error}")
            
            return {
                'success': False,
                'error': error_msg
            }

    def _customize_signature_content(self, signature_html: str, project, custom_message: str = None) -> str:
        """
        Customize signature content by replacing placeholders with project data.
        
        Args:
            signature_html: Original signature HTML from Outlook
            project: Django Project model instance
            custom_message: Optional custom message to replace CCC placeholder
            
        Returns:
            Customized signature HTML
        """
        try:
            if not signature_html:
                return signature_html
            
            # Generate timestamp for filenames
            timestamp = self._generate_timestamp()
            
            # Get film numbers
            film_numbers = self._get_project_roll_numbers(project)
            
            # Get archive ID
            archive_id = str(project.archive_id) if hasattr(project, 'archive_id') and project.archive_id else 'UNKNOWN'
            
            # Replace placeholders in signature
            customized_html = signature_html
            
            # Replace XXX with archive_id (appears in multiple places)
            customized_html = customized_html.replace('XXX', archive_id)
            
            # Replace YYY with film numbers
            if film_numbers:
                customized_html = customized_html.replace('YYY', film_numbers)
            else:
                customized_html = customized_html.replace('YYY', 'N/A')
            
            # Replace DDMMYYHHMM with actual timestamp
            customized_html = customized_html.replace('DDMMYYHHMM', timestamp)
                        
            # Replace CCC with custom message or em dash if empty
            if custom_message and custom_message.strip():
                # Convert line breaks to HTML breaks for proper display
                formatted_message = custom_message.strip().replace('\n', '<br>')
                customized_html = customized_html.replace('CCC', formatted_message)
            else:
                # Replace CCC placeholder with em dash if no custom message
                customized_html = customized_html.replace('CCC', '—')
            
            self.logger.debug(f"Customized signature placeholders: XXX -> {archive_id}, YYY -> {film_numbers or 'N/A'}, DDMMYYHHMM -> {timestamp}, QQQ -> (removed), CCC -> {custom_message[:30] if custom_message else 'em dash'}...")
            
            return customized_html
            
        except Exception as e:
            self.logger.error(f"Error customizing signature content: {e}")
            return signature_html  # Return original if customization fails

    def _generate_timestamp(self) -> str:
        """
        Generate timestamp in DDMMYYHHMM format for filenames.
        
        Returns:
            Formatted timestamp string
        """
        return datetime.now().strftime('%d%m%y%H%M')
    
    def _generate_email_body(self, project, file_paths: Dict[str, Any]) -> str:
        """
        Generate email body using Django template.
        
        Args:
            project: Django Project model instance
            file_paths: Dictionary with file paths and metadata
            
        Returns:
            Rendered HTML email body
        """
        try:
            # Get film numbers from project
            film_numbers = self._get_project_roll_numbers(project)
            
            # Prepare template context
            context = {
                'project': project,
                'film_numbers': film_numbers,
                'file_paths': file_paths,
                'generated_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            
            # Render template using the original professional signature template
            email_body = render_to_string('microapp/handoff/email/signature_template.html', context)
            return email_body
            
        except Exception as e:
            self.logger.error(f"Error generating email body: {e}")
            # Fallback to simple text
            return self._generate_default_email_body(project, file_paths)
    
    def _get_project_roll_numbers(self, project) -> str:
        """
        Get formatted film numbers for the project.
        
        Args:
            project: Django Project model instance
            
        Returns:
            Comma-separated string of film numbers
        """
        try:
            # Get all rolls for this project, ordered by film_number or roll_number
            rolls = project.rolls.all().order_by('film_number', 'roll_number')
            
            # Prefer film_number, fallback to roll_number
            film_numbers = []
            for roll in rolls:
                if roll.film_number:
                    film_numbers.append(roll.film_number)
                elif roll.roll_number:
                    film_numbers.append(roll.roll_number)
            
            if film_numbers:
                return ', '.join(film_numbers)
            else:
                return 'N/A'
                
        except Exception as e:
            self.logger.error(f"Error getting film numbers for project {project.archive_id}: {e}")
            return 'N/A'
    
    def _generate_default_email_body(self, project, file_paths: Dict[str, Any]) -> str:
        """
        Generate default email body for handoff emails.
        
        Args:
            project: Django Project model instance
            file_paths: Dictionary with file paths and metadata
            
        Returns:
            Default email body text
        """
        film_numbers = self._get_project_roll_numbers(project)
        
        return f"""Dear Team,

Please find attached the final index files for the completed microfilm project.

Project Details:
- Archive ID: {project.archive_id}
- Project Name: {project.name or 'N/A'}
- Document Type: {project.doc_type or 'Documents'}
- Location: {project.location or 'N/A'}
- Film Numbers: {film_numbers}
- Total Documents: {file_paths.get('entries_count', 'N/A')}
- Generation Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

Attached Files:
- blips.xlsx (Excel format index - {file_paths.get('excel_size', 0)} bytes)

The project has been successfully filmed, developed, validated, and is ready for final processing.

Best regards,
Iron Mountain Microfilm Team

---
This email was generated automatically by the Microfilm Processing System.
"""

    def _create_handoff_record(self, project, email_data: Dict[str, Any], file_paths: Dict[str, Any], user: User = None) -> 'HandoffRecord':
        """
        Create a handoff record to track the email being sent.
        
        Args:
            project: Django Project model instance
            email_data: Dictionary with email form data
            file_paths: Dictionary with file paths from generate_handoff_files
            user: User who initiated the handoff
            
        Returns:
            HandoffRecord instance
        """
        try:
            # Import here to avoid circular imports
            from ..models import HandoffRecord, HandoffValidationSnapshot
            
            # Generate unique handoff ID
            handoff_id = f"HO-{project.archive_id}-{datetime.now().strftime('%Y%m%d%H%M%S')}-{str(uuid.uuid4())[:8]}"
            
            # Create handoff record
            handoff_record = HandoffRecord.objects.create(
                project=project,
                user=user,
                handoff_id=handoff_id,
                recipient_email=email_data.get('to', ''),
                recipient_name=self._extract_recipient_name(email_data.get('to', '')),
                subject=email_data.get('subject', ''),
                custom_message=email_data.get('custom_message', ''),
                status='pending'
            )
            
            # Update validation summary if available
            validation_results = getattr(self, '_last_validation_results', None)
            if validation_results:
                handoff_record.update_validation_summary(validation_results)
            
            # Update file information (Excel only - DAT file disabled)
            if file_paths:
                handoff_record.update_file_info(
                    excel_path=file_paths.get('excel_path'),
                    # dat_path=file_paths.get('dat_path')  # DAT file disabled
                )
            
            # Set film numbers
            film_numbers = email_data.get('film_numbers', '')
            if not film_numbers:
                film_numbers = self._get_project_roll_numbers(project)
            
            if film_numbers:
                handoff_record.film_numbers = film_numbers
                # Count rolls
                handoff_record.total_rolls = len([fn.strip() for fn in film_numbers.split(',') if fn.strip()])
            
            handoff_record.save()
            
            # Create validation snapshots if validation data is available
            validation_data = getattr(self, '_last_validation_data', None)
            if validation_data:
                self._create_validation_snapshots(handoff_record, validation_data)
            
            self.logger.info(f"Created handoff record {handoff_id} for project {project.archive_id}")
            return handoff_record
            
        except Exception as e:
            self.logger.error(f"Failed to create handoff record for project {project.archive_id}: {e}")
            # Create a minimal record to avoid breaking the email process
            try:
                from ..models import HandoffRecord
                return HandoffRecord.objects.create(
                    project=project,
                    user=user,
                    handoff_id=f"HO-{project.archive_id}-{datetime.now().strftime('%Y%m%d%H%M%S')}-ERROR",
                    recipient_email=email_data.get('to', ''),
                    subject=email_data.get('subject', ''),
                    status='pending'
                )
            except Exception as fallback_error:
                self.logger.error(f"Failed to create fallback handoff record: {fallback_error}")
                raise e
    
    def _extract_recipient_name(self, email_address: str) -> str:
        """
        Extract recipient name from email address.
        
        Args:
            email_address: Email address string
            
        Returns:
            Extracted name or empty string
        """
        try:
            # Handle format like "Name <email@domain.com>"
            if '<' in email_address and '>' in email_address:
                name_part = email_address.split('<')[0].strip()
                return name_part.strip('"').strip("'")
            
            # Handle format like "name@domain.com"
            # Extract name from email local part
            local_part = email_address.split('@')[0]
            # Convert dots and underscores to spaces and title case
            name = local_part.replace('.', ' ').replace('_', ' ').title()
            return name
            
        except Exception:
            return ''
    
    def _create_validation_snapshots(self, handoff_record: 'HandoffRecord', validation_data: List[Dict]) -> None:
        """
        Create validation snapshots for audit purposes.
        
        Args:
            handoff_record: HandoffRecord instance
            validation_data: List of validation data dictionaries
        """
        try:
            from ..models import HandoffValidationSnapshot
            
            snapshots = []
            for item in validation_data:
                # Determine validation status and issues
                missing_com_id = not item.get('com_id') or item.get('com_id') == ''
                missing_film_blip = not item.get('film_blip') or item.get('film_blip') == ''
                
                # Determine overall status
                if missing_com_id:
                    status = 'error'
                    message = 'Missing COM ID'
                elif missing_film_blip:
                    status = 'pending'
                    message = 'Film blip not found in logs'
                elif item.get('status') == 'validated':
                    status = 'validated'
                    message = 'Validation successful'
                elif item.get('status') == 'warning':
                    status = 'warning'
                    message = item.get('message', 'Minor validation issue')
                else:
                    status = 'error'
                    message = item.get('message', 'Validation failed')
                
                snapshot = HandoffValidationSnapshot(
                    handoff_record=handoff_record,
                    document_id=item.get('document_id', ''),
                    roll_number=item.get('roll', ''),
                    barcode=item.get('barcode', ''),
                    com_id=str(item.get('com_id', '')) if item.get('com_id') else None,
                    temp_blip=item.get('temp_blip', ''),
                    film_blip=item.get('film_blip', ''),
                    validation_status=status,
                    validation_message=message,
                    missing_com_id=missing_com_id,
                    missing_film_blip=missing_film_blip,
                    blip_mismatch=(item.get('temp_blip') != item.get('film_blip')) if item.get('temp_blip') and item.get('film_blip') else False
                )
                snapshots.append(snapshot)
            
            # Bulk create snapshots
            HandoffValidationSnapshot.objects.bulk_create(snapshots)
            self.logger.info(f"Created {len(snapshots)} validation snapshots for handoff {handoff_record.handoff_id}")
            
        except Exception as e:
            self.logger.error(f"Failed to create validation snapshots: {e}")

    def store_validation_results(self, validation_results: Dict, validation_data: List[Dict]) -> None:
        """
        Store validation results for later use in handoff record creation.
        
        Args:
            validation_results: Validation summary results
            validation_data: List of validation data dictionaries
        """
        self._last_validation_results = validation_results
        self._last_validation_data = validation_data

 